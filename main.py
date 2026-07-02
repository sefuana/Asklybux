#!/usr/bin/env python3
"""
AX Worker Bot
Production-ready Telegram bot for Railway + GitHub deployment.
"""

import asyncio
import json
import logging
import os
import random
import re
import string
import tempfile
from datetime import datetime, timezone
import hashlib
import hmac
import base64
import struct
import time
import firebase_admin
import openpyxl
from firebase_admin import credentials, db
from telegram import (
    InlineKeyboardButton,
    InlineKeyboardMarkup,
    ReplyKeyboardMarkup,
    Update,
)
from telegram.ext import (
    Application,
    CallbackQueryHandler,
    CommandHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)

# ─────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────
logging.basicConfig(
    format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# Environment Variables
# ─────────────────────────────────────────────
BOT_TOKEN = os.environ["BOT_TOKEN"]
FIREBASE_CONFIG = os.environ["FIREBASE_CONFIG"]
REQUIRED_CHANNEL = os.environ.get("REQUIRED_CHANNEL", "")
CHANNEL_LINK = os.environ.get("CHANNEL_LINK", "")
FIREBASE_DATABASE_URL = os.environ["FIREBASE_DATABASE_URL"]

# ─────────────────────────────────────────────
# Admin IDs
# ─────────────────────────────────────────────
ADMIN_IDS = {8907284640, 8760645843}

_pending_auto_approve: dict = {}
USD_TO_BDT_RATE = 120.0
BOT_ENABLED = True

# ─────────────────────────────────────────────
# Conversation States
# ─────────────────────────────────────────────
(
    HOME,
    TASK_MENU,
    TASK_2FA_INFO,
    TASK_2FA_STARTED,
    TASK_2FA_AWAIT_KEY,
    TASK_2FA_1H_INFO,
    TASK_2FA_1H_STARTED,
    TASK_2FA_1H_AWAIT_KEY,
    WITHDRAW_MENU,
    WITHDRAW_AMOUNT,
    WITHDRAW_ADDRESS,
    ADMIN_ACTS_VIEW,
    TASK_FB_INFO,
    TASK_FB_AWAIT_UID,
    TASK_FB_AWAIT_COOKIES,
    TASK_FB_STARTED,
    ADMIN_FBACTS_VIEW,
    SETTINGS_MENU,
    WITHDRAW_CONFIRM,
) = range(19)

# ─────────────────────────────────────────────
# Firebase Initialisation
# ─────────────────────────────────────────────
def init_firebase():
    config = json.loads(FIREBASE_CONFIG)
    config.pop("databaseURL", None)
    database_url = FIREBASE_DATABASE_URL.strip()
    if not database_url:
        raise ValueError("FIREBASE_DATABASE_URL environment variable is missing or empty.")
    cred = credentials.Certificate(config)
    firebase_admin.initialize_app(cred, {"databaseURL": database_url})
    logger.info("Firebase initialised successfully.")

# ─────────────────────────────────────────────
# UI/UX Helpers
# ─────────────────────────────────────────────
DIV = "━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
DIV_SHORT = "━━━━━━━━━━━━━━━━━━"

def section(title: str, body: str, footer: str = "") -> str:
    """Build a consistently formatted message block."""
    parts = [DIV, f"  {title}", DIV, "", body]
    if footer:
        parts += ["", DIV, footer]
    else:
        parts.append(DIV)
    return "\n".join(parts)

def fmt_usd(amount: float) -> str:
    return f"${amount:.4f}"

def fmt_bdt(amount: float) -> str:
    return f"৳{amount:.2f}"

# ─────────────────────────────────────────────
# Referral Helpers
# ─────────────────────────────────────────────
def get_referral_data(user_id: int) -> dict:
    try:
        data = db.reference(f"referrals/{user_id}").get()
        if data is None:
            data = {
                "referral_code": str(user_id),
                "referred_by": None,
                "referrals": [],
                "total_earned": 0.0,
                "joined_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
            }
            db.reference(f"referrals/{user_id}").set(data)
        return data
    except Exception as e:
        logger.error(f"get_referral_data({user_id}) failed: {e}")
        return {"referral_code": str(user_id), "referred_by": None, "referrals": [], "total_earned": 0.0}


def update_referral_data(user_id: int, updates: dict):
    try:
        db.reference(f"referrals/{user_id}").update(updates)
    except Exception as e:
        logger.error(f"update_referral_data({user_id}) failed: {e}")


def add_referral(user_id: int, referrer_id: int):
    if user_id == referrer_id:
        return False
    try:
        referrer_data = get_referral_data(referrer_id)
        user_data = get_referral_data(user_id)
        if user_data.get("referred_by") is not None:
            return False
        update_referral_data(user_id, {"referred_by": referrer_id})
        referrals = referrer_data.get("referrals", [])
        if user_id not in referrals:
            referrals.append(user_id)
            update_referral_data(referrer_id, {"referrals": referrals})
        logger.info(f"Referral added: {user_id} referred by {referrer_id}")
        return True
    except Exception as e:
        logger.error(f"add_referral({user_id}, {referrer_id}) failed: {e}")
        return False


def get_referral_stats(user_id: int) -> dict:
    data = get_referral_data(user_id)
    referrals = data.get("referrals", [])
    now = datetime.now(timezone.utc)
    new_last_24h = 0
    for ref_id in referrals:
        ref_data = get_referral_data(ref_id)
        joined_at_str = ref_data.get("joined_at", "")
        if joined_at_str:
            try:
                joined_at = datetime.strptime(joined_at_str, "%Y-%m-%d %H:%M:%S UTC").replace(tzinfo=timezone.utc)
                if (now - joined_at).total_seconds() <= 24 * 3600:
                    new_last_24h += 1
            except Exception:
                pass
    return {
        "total": len(referrals),
        "new_last_24h": new_last_24h,
        "total_earned": data.get("total_earned", 0.0)
    }

# ─────────────────────────────────────────────
# Bot State
# ─────────────────────────────────────────────
def get_bot_state() -> dict:
    try:
        state = db.reference("bot/state").get()
        if state is None:
            state = {
                "enabled": True,
                "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                "updated_by": None
            }
            db.reference("bot/state").set(state)
        return state
    except Exception as e:
        logger.error(f"get_bot_state failed: {e}")
        return {"enabled": True, "last_updated": "", "updated_by": None}


def set_bot_state(enabled: bool, admin_id: int = None):
    try:
        state = {
            "enabled": enabled,
            "last_updated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "updated_by": admin_id
        }
        db.reference("bot/state").set(state)
        logger.info(f"Bot state changed to {'ON' if enabled else 'OFF'} by admin {admin_id}")
    except Exception as e:
        logger.error(f"set_bot_state failed: {e}")

# ─────────────────────────────────────────────
# Leaderboard Helpers
# ─────────────────────────────────────────────
def get_today_stats() -> dict:
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    all_subs = get_all_submissions()
    today_stats = {}
    for uid, subs in all_subs.items():
        if subs:
            today_count = 0
            for sub_id, sub_data in subs.items():
                sub_date = sub_data.get('datetime', '')
                if sub_date.startswith(today):
                    today_count += 1
            if today_count > 0:
                today_stats[uid] = today_count
    return today_stats


def get_leaderboard_data() -> dict:
    try:
        data = db.reference("leaderboard/data").get()
        return data or {}
    except Exception as e:
        logger.error(f"get_leaderboard_data failed: {e}")
        return {}


def set_leaderboard_data(data: dict):
    try:
        db.reference("leaderboard/data").set(data)
    except Exception as e:
        logger.error(f"set_leaderboard_data failed: {e}")


def get_leaderboard_settings() -> dict:
    try:
        settings = db.reference("leaderboard/settings").get()
        if settings is None:
            settings = {
                "mode": "auto",
                "last_update": "",
                "current_prizes": {},
                "enabled": True
            }
            db.reference("leaderboard/settings").set(settings)
        return settings
    except Exception as e:
        logger.error(f"get_leaderboard_settings failed: {e}")
        return {"mode": "auto", "last_update": "", "current_prizes": {}, "enabled": True}


def set_leaderboard_settings(updates: dict):
    try:
        existing = db.reference("leaderboard/settings").get() or {}
        existing.update(updates)
        db.reference("leaderboard/settings").set(existing)
    except Exception as e:
        logger.error(f"set_leaderboard_settings failed: {e}")

# ─────────────────────────────────────────────
# Leaderboard Generation
# ─────────────────────────────────────────────
def generate_real_leaderboard() -> dict:
    today_stats = get_today_stats()
    sorted_users = sorted(today_stats.items(), key=lambda x: x[1], reverse=True)[:10]
    prizes = [2.0, 1.0, 0.5, 0.5, 0.5, 0.2, 0.2, 0.2, 0.2, 0.1]
    leaderboard = {}
    for idx, (uid, count) in enumerate(sorted_users):
        if idx < len(prizes):
            leaderboard[uid] = {
                "completed": count,
                "prize": prizes[idx],
                "rank": idx + 1
            }
    return leaderboard


def generate_auto_leaderboard(increment: int = 0, previous_leaderboard: dict = None) -> dict:
    if increment and previous_leaderboard:
        leaderboard = {}
        uid_list = []
        counts = []
        for uid, data in previous_leaderboard.items():
            old_count = data.get('completed', 0)
            increment_amount = random.randint(3, 5)
            new_count = old_count + increment_amount
            uid_list.append(uid)
            counts.append(new_count)
        sorted_pairs = sorted(zip(counts, uid_list), reverse=True)
        counts = [c for c, _ in sorted_pairs]
        uid_list = [u for _, u in sorted_pairs]
        prizes = [4.0, 2.0, 1.0, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5]
        for idx, (uid, count) in enumerate(zip(uid_list, counts)):
            if idx < len(prizes):
                leaderboard[uid] = {
                    "completed": count,
                    "prize": prizes[idx],
                    "rank": idx + 1,
                    "masked_id": str(uid)[:4] + "***" + str(uid)[-2:] if len(str(uid)) > 6 else str(uid)
                }
        return leaderboard

    prizes = [4.0, 2.0, 1.0, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5]
    used_ids = set()
    uid_list = []
    while len(uid_list) < 10:
        uid = random.randint(100000000, 999999999)
        if uid not in used_ids:
            used_ids.add(uid)
            uid_list.append(str(uid))
    counts = [random.randint(10, 100) for _ in range(10)]
    counts.sort(reverse=True)
    leaderboard = {}
    for idx, (uid, count) in enumerate(zip(uid_list, counts)):
        leaderboard[uid] = {
            "completed": count,
            "prize": prizes[idx],
            "rank": idx + 1,
            "masked_id": str(uid)[:4] + "***" + str(uid)[-2:] if len(str(uid)) > 6 else str(uid)
        }
    return leaderboard


RANK_MEDALS = {1: "🥇", 2: "🥈", 3: "🥉"}


def format_leaderboard_text(leaderboard: dict, mode: str) -> str:
    if not leaderboard:
        return (
            f"{DIV}\n"
            "  🏆  Leaderboard\n"
            f"{DIV}\n\n"
            "No data available yet.\n"
            "Check back after the next update.\n\n"
            f"{DIV}"
        )

    lines = [
        DIV,
        "  🏆  Daily Leaderboard",
        DIV,
        "",
        "Results announced daily at 01:00 Helsinki time.",
        "Top performers receive real balance rewards!",
        "",
        "🔄  Updates every day at 1 PM GMT+6",
        "",
        DIV_SHORT,
        "  Rank   User ID        Tasks     Prize",
        DIV_SHORT,
    ]

    for uid, data in sorted(leaderboard.items(), key=lambda x: x[1]['rank']):
        rank = data.get('rank', 0)
        completed = data.get('completed', 0)
        prize = data.get('prize', 0)

        if len(str(uid)) > 6:
            masked = str(uid)[:4] + "***" + str(uid)[-2:]
        else:
            masked = str(uid)

        medal = RANK_MEDALS.get(rank, f" {rank}.")

        if mode == "auto":
            task_text = f"{completed} tasks"
        else:
            task_text = f"{completed}"

        lines.append(f"{medal}  {masked:<14} {task_text:<10} ${prize}")

    lines += [
        DIV_SHORT,
        "",
        "Complete more tasks today to climb the ranks!",
        DIV,
    ]
    return "\n".join(lines)

# ─────────────────────────────────────────────
# Firebase Helpers
# ─────────────────────────────────────────────
def get_user(user_id: int) -> dict:
    try:
        ref = db.reference(f"users/{user_id}")
        data = ref.get()
        if data is None:
            data = {
                "balance": 0.0,
                "approved": 0,
                "in_review": 0,
                "total_submitted": 0,
            }
            ref.set(data)
        return data
    except Exception as e:
        logger.error(f"get_user({user_id}) failed: {e}")
        return {"balance": 0.0, "approved": 0, "in_review": 0, "total_submitted": 0}


def update_user(user_id: int, updates: dict):
    try:
        db.reference(f"users/{user_id}").update(updates)
    except Exception as e:
        logger.error(f"update_user({user_id}) failed: {e}")


def add_submission(user_id: int, tg_username: str, username: str, password: str, key: str):
    try:
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        sub_ref = db.reference(f"submissions/{user_id}").push()
        sub_ref.set({
            "username": username,
            "password": password,
            "key": key,
            "tg_username": tg_username,
            "user_id": str(user_id),
            "datetime": now,
            "status": "pending",
        })
        user = get_user(user_id)
        update_user(user_id, {
            "in_review": user.get("in_review", 0) + 1,
            "total_submitted": user.get("total_submitted", 0) + 1,
        })
        _rebuild_xlsx(user_id)
    except Exception as e:
        logger.error(f"add_submission({user_id}) failed: {e}")


def get_submissions(user_id: int) -> list:
    try:
        data = db.reference(f"submissions/{user_id}").get()
        if not data:
            return []
        return [
            {"id": k, **v} for k, v in data.items()
            if v.get("status", "pending") == "pending"
        ]
    except Exception as e:
        logger.error(f"get_submissions({user_id}) failed: {e}")
        return []


def remove_submissions(user_id: int):
    try:
        count = len(get_submissions(user_id))
        db.reference(f"submissions/{user_id}").delete()
        try:
            db.reference(f"xlsx_cache/{user_id}").delete()
        except Exception:
            pass
        user = get_user(user_id)
        new_review = max(0, user.get("in_review", 0) - count)
        update_user(user_id, {"in_review": new_review})
    except Exception as e:
        logger.error(f"remove_submissions({user_id}) failed: {e}")


def create_withdrawal(user_id: int, tg_username: str, amount: float, wallet: str) -> str:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    ref = db.reference(f"withdrawals/{user_id}").push()
    fee = 0.025
    receive = round(amount - fee, 4)
    ref.set({
        "tg_username": tg_username,
        "amount": amount,
        "fee": fee,
        "receive": receive,
        "wallet": wallet,
        "status": "pending",
        "datetime": now,
    })
    return ref.key


def get_withdrawal(user_id: int, w_id: str):
    try:
        return db.reference(f"withdrawals/{user_id}/{w_id}").get()
    except Exception as e:
        logger.error(f"get_withdrawal({user_id}, {w_id}) failed: {e}")
        return None


def update_withdrawal(user_id: int, w_id: str, updates: dict):
    try:
        db.reference(f"withdrawals/{user_id}/{w_id}").update(updates)
    except Exception as e:
        logger.error(f"update_withdrawal({user_id}, {w_id}) failed: {e}")


def get_all_users() -> dict:
    try:
        data = db.reference("users").get()
        return data or {}
    except Exception as e:
        logger.error(f"get_all_users() failed: {e}")
        return {}


def get_all_submissions() -> dict:
    try:
        data = db.reference("submissions").get()
        return data or {}
    except Exception as e:
        logger.error(f"get_all_submissions() failed: {e}")
        return {}


def get_pending_withdrawal_total(user_id: int) -> float:
    """Return the sum of all pending withdrawal amounts for a user."""
    try:
        withdrawals = db.reference(f"withdrawals/{user_id}").get()
        if not withdrawals:
            return 0.0
        total = 0.0
        for w in withdrawals.values():
            if isinstance(w, dict) and w.get("status") == "pending":
                total += w.get("amount", 0.0)
        return round(total, 4)
    except Exception as e:
        logger.error(f"get_pending_withdrawal_total({user_id}) failed: {e}")
        return 0.0


def get_all_pending_withdrawals() -> list:
    """Return a flat list of all pending withdrawals across all users, sorted newest first."""
    try:
        all_wds = db.reference("withdrawals").get() or {}
        result = []
        for uid_str, wds in all_wds.items():
            if not isinstance(wds, dict):
                continue
            for w_id, w_data in wds.items():
                if not isinstance(w_data, dict):
                    continue
                if w_data.get("status") == "pending":
                    result.append({
                        "user_id": int(uid_str),
                        "w_id": w_id,
                        **w_data,
                    })
        result.sort(key=lambda x: x.get("datetime", ""), reverse=True)
        return result
    except Exception as e:
        logger.error(f"get_all_pending_withdrawals() failed: {e}")
        return []

# ─────────────────────────────────────────────
# Task Price Settings
# ─────────────────────────────────────────────
def get_task_price() -> float:
    try:
        price = db.reference("settings/task_price").get()
        if price is None:
            price = 0.0330
            db.reference("settings/task_price").set(price)
        return float(price)
    except Exception as e:
        logger.error(f"get_task_price failed: {e}")
        return 0.0330


def set_task_price(price: float):
    try:
        db.reference("settings/task_price").set(round(price, 4))
    except Exception as e:
        logger.error(f"set_task_price failed: {e}")


def get_task_1h_price() -> float:
    try:
        price = db.reference("settings/task_1h_price").get()
        if price is None:
            price = 0.220
            db.reference("settings/task_1h_price").set(price)
        return float(price)
    except Exception as e:
        logger.error(f"get_task_1h_price failed: {e}")
        return 0.220


def set_task_1h_price(price: float):
    try:
        db.reference("settings/task_1h_price").set(round(price, 4))
    except Exception as e:
        logger.error(f"set_task_1h_price failed: {e}")


def get_task_settings() -> dict:
    try:
        raw_6h = db.reference("settings/tasks/task_6h_enabled").get()
        raw_1h = db.reference("settings/tasks/task_1h_enabled").get()
        last_updated = db.reference("settings/tasks/last_updated").get() or ""

        def parse_bool(val, default=True):
            if val is None:
                return default
            if isinstance(val, bool):
                return val
            if isinstance(val, str):
                return val.lower() != "false"
            return bool(val)

        return {
            "task_6h_enabled": parse_bool(raw_6h, True),
            "task_1h_enabled": parse_bool(raw_1h, True),
            "last_updated": last_updated,
        }
    except Exception as e:
        logger.error(f"get_task_settings failed: {e}")
        return {"task_6h_enabled": True, "task_1h_enabled": True, "last_updated": ""}


def set_task_settings(updates: dict):
    try:
        existing = db.reference("settings/tasks").get() or {}
        existing.update(updates)
        existing["last_updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        db.reference("settings/tasks").set(existing)
    except Exception as e:
        logger.error(f"set_task_settings failed: {e}")

# ─────────────────────────────────────────────
# XLSX Helpers
# ─────────────────────────────────────────────
XLSX_DIR = tempfile.gettempdir()


def _xlsx_path(user_id: int) -> str:
    return os.path.join(XLSX_DIR, f"submissions_{user_id}.xlsx")


def _rebuild_xlsx(user_id: int):
    try:
        subs = get_submissions(user_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Submissions"
        ws.append(["Username", "Password", "2FA Key", "TG Username", "User ID", "DateTime"])
        for row in subs:
            ws.append([
                row.get("username", ""),
                row.get("password", ""),
                row.get("key", ""),
                row.get("tg_username", ""),
                row.get("user_id", str(user_id)),
                row.get("datetime", ""),
            ])
        path = _xlsx_path(user_id)
        wb.save(path)
        logger.info(f"XLSX rebuilt for user {user_id} → {path}")
    except Exception as e:
        logger.error(f"_rebuild_xlsx({user_id}) failed: {e}")


def build_xlsx_bytes(user_id: int) -> bytes:
    path = _xlsx_path(user_id)
    if not os.path.exists(path):
        _rebuild_xlsx(user_id)
    try:
        with open(path, "rb") as f:
            return f.read()
    except Exception as e:
        logger.error(f"build_xlsx_bytes({user_id}) failed: {e}")
        return b""

# ─────────────────────────────────────────────
# Utility Functions
# ─────────────────────────────────────────────
UNCOMMON_ADJECTIVES = [
    "lunar", "crimson", "azure", "phantom", "mystic", "neon", "velvet",
    "cobalt", "sable", "ivory", "scarlet", "gilded", "obsidian", "spectral",
    "vivid", "prism", "hollow", "ancient", "frosted", "runic",
]
UNCOMMON_NOUNS = [
    "wraith", "oracle", "cipher", "nexus", "specter", "herald", "vortex",
    "relic", "sigil", "golem", "comet", "prism", "mirage", "scion",
    "phantom", "herald", "bastion", "ember", "zenith", "epoch",
]

_HEX_CHARS = "0123456789abcdef"


def generate_username() -> str:
    adj = random.choice(UNCOMMON_ADJECTIVES)
    noun = random.choice(UNCOMMON_NOUNS)
    num = random.randint(100, 9999)
    return f"{adj}_{noun}_{num}"


def generate_tx_hash() -> str:
    return "0x" + "".join(random.choices(_HEX_CHARS, k=64))


_BASE32_RE = re.compile(r"^[A-Z2-7]{16,32}$")
_BEP20_RE = re.compile(r"^0x[0-9a-fA-F]{40}$")  # এই লাইনটা যোগ করুন


def get_default_password() -> str:
    try:
        password = db.reference("settings/default_password").get()
        if password is None:
            password = "axiex@25"
            db.reference("settings/default_password").set(password)
        return password
    except Exception as e:
        logger.error(f"get_default_password failed: {e}")
        return "axiex@25"


def is_valid_bep20(address: str) -> bool:
    return bool(_BEP20_RE.match(address))
    
def validate_2fa_key(raw: str) -> tuple:
    original = raw.strip()
    cleaned = original.replace(" ", "").upper()
    if len(cleaned) not in (16, 32):
        return None, None, (
            f"❌ Invalid key length. Expected 16 or 32 characters — "
            f"you entered {len(cleaned)} (spaces excluded). Please try again:"
        )
    if not _BASE32_RE.match(cleaned):
        return None, None, (
            "❌ Invalid key format. Only letters A–Z and digits 2–7 are allowed. Please try again:"
        )
    return cleaned, original, None


# ─────────────────────────────────────────────
# TOTP Generator
# ─────────────────────────────────────────────
def generate_totp(secret_key: str, interval: int = 30) -> str:
    key_bytes = base64.b32decode(secret_key, casefold=True)
    current_time = int(time.time())
    time_step = current_time // interval
    time_bytes = struct.pack(">Q", time_step)
    hmac_hash = hmac.new(key_bytes, time_bytes, hashlib.sha1).digest()
    offset = hmac_hash[-1] & 0x0F
    code_bytes = hmac_hash[offset:offset + 4]
    otp = struct.unpack(">I", code_bytes)[0] & 0x7FFFFFFF
    otp = otp % 1000000
    return f"{otp:06d}"
    
def get_all_fb_users() -> dict:
    """Get all Facebook users from Firebase."""
    try:
        data = db.reference("fb_users").get()
        return data or {}
    except Exception as e:
        logger.error(f"get_all_fb_users() failed: {e}")
        return {}
        
def build_fb_xlsx_bytes(user_id: int) -> bytes:
    """Build XLSX file for FB submissions."""
    try:
        subs = get_fb_submissions(user_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FB Submissions"
        ws.append(["UID", "Password", "First Name", "Last Name", "TG Username", "User ID", "DateTime"])
        for row in subs:
            ws.append([
                row.get("uid", ""),
                row.get("password", ""),
                row.get("firstname", ""),
                row.get("lastname", ""),
                row.get("tg_username", ""),
                row.get("user_id", str(user_id)),
                row.get("datetime", ""),
            ])
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            with open(tmp.name, "rb") as f:
                data = f.read()
            os.remove(tmp.name)
            return data
    except Exception as e:
        logger.error(f"build_fb_xlsx_bytes({user_id}) failed: {e}")
        return b""
        
def approve_fb_submission(user_id: int, submission_id: str, task_price: float) -> bool:
    """Approve a Facebook submission and credit user."""
    try:
        sub_ref = db.reference(f"fb_submissions/{user_id}/{submission_id}")
        submission = sub_ref.get()
        if not submission:
            return False
        
        fb_user = get_fb_user(user_id)
        new_approved = fb_user.get("approved", 0) + 1
        new_balance = round(fb_user.get("balance", 0.0) + task_price, 4)
        new_in_review = max(0, fb_user.get("in_review", 0) - 1)
        
        update_fb_user(user_id, {
            "approved": new_approved,
            "balance": new_balance,
            "in_review": new_in_review
        })
        
        sub_ref.delete()  # Remove after approval
        return True
    except Exception as e:
        logger.error(f"approve_fb_submission failed: {e}")
        return False

# ─────────────────────────────────────────────
# Keyboards
# ─────────────────────────────────────────────
HOME_KEYBOARD = ReplyKeyboardMarkup(
    [
        ["💼 Dashboard"],
        ["💰 Wallet", "📋 Tasks"],
        ["👥 Invite Friends", "🏆 Leaderboard"],
        ["📥 Withdraw", "👤 Profile"],
        ["📞 Support"],
    ],
    resize_keyboard=True,
)

TASK_MENU_KEYBOARD = ReplyKeyboardMarkup(
    [
        ["📱 Instagram 2FA — $0.0300"],
        ["⭐ Instagram 2FA Premium — $0.220"],
        ["🍪 Facebook Cookie — $0.0350"],
        ["🔙 Back"],
    ],
    resize_keyboard=True,
)

TASK_START_KEYBOARD = ReplyKeyboardMarkup(
    [["▶️ Start Task"], ["🔙 Cancel"]],
    resize_keyboard=True,
)

WITHDRAW_NEW_KEYBOARD = ReplyKeyboardMarkup(
    [["💎 USDT-BEP20"], ["Cancel ❌"]],
    resize_keyboard=True,
)

WITHDRAW_MENU_KEYBOARD = ReplyKeyboardMarkup(
    [["💎 USDT — BEP20"], ["📱 bKash — BDT"], ["🔙 Back"]],
    resize_keyboard=True,
)

BACK_KEYBOARD = ReplyKeyboardMarkup(
    [["🔙 Back"]],
    resize_keyboard=True,
)

SETTINGS_KEYBOARD = ReplyKeyboardMarkup(
    [["ℹ️ About Bot"], ["🔙 Back"]],
    resize_keyboard=True,
)

# ─────────────────────────────────────────────
# /start
# ─────────────────────────────────────────────
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    user_id = user.id

    bot_state = get_bot_state()
    if not bot_state.get("enabled", True):
        await update.message.reply_text(
            f"{DIV}\n"
            "  🔴  Bot Offline\n"
            f"{DIV}\n\n"
            "The bot is currently under maintenance.\n\n"
            f"Since: {bot_state.get('last_updated', 'Unknown')}\n\n"
            "Please try again later.\n\n"
            "For urgent matters, reach us at:\n"
            "  @axWorker_Admin\n\n"
            f"{DIV}"
        )
        return HOME

    joined = await check_user_joined(user_id, context)

    if not joined and REQUIRED_CHANNEL:
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📢 Join Channel", url=CHANNEL_LINK)],
            [InlineKeyboardButton("✅ I've Joined — Check", callback_data="check_join")]
        ])
        await update.message.reply_text(
            f"{DIV}\n"
            "  🔒  Access Restricted\n"
            f"{DIV}\n\n"
            f"Welcome, {user.first_name or 'there'}!\n\n"
            "To use this bot, you must first join\n"
            "our official channel.\n\n"
            "👇 Join below, then tap Check.\n\n"
            f"{DIV}",
            reply_markup=keyboard
        )
        return HOME

    return await _start_bot(update, context)


async def _start_bot(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    user_id = user.id

    if context.args and len(context.args) > 0:
        referrer_id_str = context.args[0]
        try:
            referrer_id = int(referrer_id_str)
            if referrer_id != user_id:
                user_ref_data = get_referral_data(user_id)
                if user_ref_data.get("referred_by") is None:
                    add_referral(user_id, referrer_id)
                    await update.message.reply_text(
                        f"{DIV}\n"
                        "  🎉  Welcome Gift\n"
                        f"{DIV}\n\n"
                        "You were referred by an existing member!\n\n"
                        "Start completing tasks to earn money.\n"
                        "Your referrer earns a bonus for every\n"
                        "task you complete. 💰\n\n"
                        f"{DIV}"
                    )
        except ValueError:
            pass

    try:
        get_user(user.id)
        get_referral_data(user.id)
    except Exception as e:
        logger.error(f"cmd_start get_user failed: {e}")

    full_name = user.full_name or user.first_name or "User"
    username = f"@{user.username}" if user.username else "No username"
    mention = f"[{full_name}](tg://user?id={user_id})"

    admin_notification = (
        f"🔔 New User\n\n"
        f"Name: {mention}\n"
        f"ID: `{user_id}`\n"
        f"Username: {username}\n"
        f"Time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n\n"
        f"Total users: `{len(get_all_users())}`"
    )

    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=admin_notification,
                parse_mode=None
            )
        except Exception as e:
            logger.warning(f"Could not notify admin {admin_id}: {e}")

    await update.message.reply_text(
        f"{DIV}\n"
        "  💼  Welcome Back\n"
        f"{DIV}\n\n"
        f"Hello, {full_name}! 👋\n\n"
        "Complete tasks and earn real money.\n"
        "Use the menu below to get started.\n\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD,
    )
    context.user_data.clear()
    return HOME

# ─────────────────────────────────────────────
# Channel Join Check
# ─────────────────────────────────────────────
async def check_user_joined(user_id: int, context: ContextTypes.DEFAULT_TYPE) -> bool:
    if not REQUIRED_CHANNEL:
        return True
    try:
        chat_member = await context.bot.get_chat_member(chat_id=REQUIRED_CHANNEL, user_id=user_id)
        return chat_member.status in ["member", "administrator", "creator"]
    except Exception as e:
        logger.error(f"Check join failed for {user_id}: {e}")
        return False


async def callback_check_join(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id
    await query.answer()
    joined = await check_user_joined(user_id, context)
    if joined:
        await query.edit_message_text(
            f"{DIV}\n"
            "  ✅  Verified\n"
            f"{DIV}\n\n"
            "You're now a channel member.\n"
            "Full access has been granted. 🎉\n\n"
            f"{DIV}"
        )
        user = query.from_user
        user_id = user.id
        full_name = user.full_name or user.first_name or "User"
        username = f"@{user.username}" if user.username else "No username"
        mention = f"[{full_name}](tg://user?id={user_id})"
        try:
            get_user(user_id)
            get_referral_data(user_id)
        except Exception as e:
            logger.error(f"callback_check_join get_user failed: {e}")
        admin_notification = (
            f"🔔 New User (via join check)\n\n"
            f"Name: {mention}\n"
            f"ID: `{user_id}`\n"
            f"Username: {username}\n"
            f"Time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n\n"
            f"Total users: `{len(get_all_users())}`"
        )
        for admin_id in ADMIN_IDS:
            try:
                await context.bot.send_message(
                    chat_id=admin_id,
                    text=admin_notification,
                    parse_mode=None
                )
            except Exception as e:
                logger.warning(f"Could not notify admin {admin_id}: {e}")
        await context.bot.send_message(
            chat_id=user_id,
            text=(
                f"{DIV}\n"
                "  💼  Welcome\n"
                f"{DIV}\n\n"
                f"Hello, {full_name}! 👋\n\n"
                "Complete tasks and earn real money.\n"
                "Use the menu below to get started.\n\n"
                f"{DIV}"
            ),
            reply_markup=HOME_KEYBOARD,
        )
        context.user_data.clear()
    else:
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📢 Join Channel", url=CHANNEL_LINK)],
            [InlineKeyboardButton("✅ Check Again", callback_data="check_join")]
        ])
        try:
            await query.edit_message_text(
                f"{DIV}\n"
                "  ❌  Not Joined Yet\n"
                f"{DIV}\n\n"
                "You haven't joined the channel yet.\n\n"
                "Please join and then tap Check Again.\n\n"
                f"{DIV}",
                reply_markup=keyboard
            )
        except Exception as e:
            if "Message is not modified" not in str(e):
                raise

# ─────────────────────────────────────────────
# Bot Enable Check
# ─────────────────────────────────────────────
async def check_bot_enabled(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    bot_state = get_bot_state()
    if not bot_state.get("enabled", True):
        if update.effective_user and update.effective_user.id in ADMIN_IDS:
            return True
        if update.message:
            await update.message.reply_text(
                f"{DIV}\n"
                "  🔴  Bot Offline\n"
                f"{DIV}\n\n"
                "The bot is currently under maintenance.\n\n"
                f"Since: {bot_state.get('last_updated', 'Unknown')}\n\n"
                "Please try again later.\n\n"
                "For urgent matters:\n"
                "  @axWorker_Admin\n\n"
                f"{DIV}"
            )
        return False
    return True

# ─────────────────────────────────────────────
# HOME Handlers
# ─────────────────────────────────────────────
async def handle_dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    try:
        data = get_user(user_id)
        ref_stats = get_referral_stats(user_id)
        task_price = get_task_price()
        balance = data.get("balance", 0.0)
        approved = data.get("approved", 0)
        in_review = data.get("in_review", 0)
        total_submitted = data.get("total_submitted", 0)
        total_refs = ref_stats.get("total", 0)
        ref_earned = ref_stats.get("total_earned", 0.0)
    except Exception as e:
        logger.error(f"handle_dashboard failed: {e}")
        await update.message.reply_text("⚠️ Could not load dashboard. Please try again.")
        return HOME

    await update.message.reply_text(
        f"{DIV}\n"
        "  💼  Dashboard\n"
        f"{DIV}\n\n"
        "💳  EARNINGS\n"
        f"   Balance          {fmt_usd(balance)}\n"
        f"   Approved Tasks   {approved}\n\n"
        "📋  ACTIVITY\n"
        f"   In Review        {in_review}\n"
        f"   Total Submitted  {total_submitted}\n\n"
        f"{DIV_SHORT}\n\n"
        "👥  REFERRALS\n"
        f"   Total Invites    {total_refs}\n"
        f"   Referral Earned  {fmt_usd(ref_earned)}\n\n"
        f"{DIV_SHORT}\n\n"
        "⚙️  TASK INFO\n"
        f"   Task Reward      {fmt_usd(task_price)}\n"
        f"   Referral Bonus   8% per task\n\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME


async def handle_balance(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    try:
        data = get_user(user_id)
        balance = data.get("balance", 0.0)
        approved = data.get("approved", 0)
        in_review = data.get("in_review", 0)
        total_submitted = data.get("total_submitted", 0)
    except Exception as e:
        logger.error(f"handle_balance failed: {e}")
        await update.message.reply_text("⚠️ Could not load wallet. Please try again.", reply_markup=HOME_KEYBOARD)
        return HOME

    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("📥 Withdraw Funds", callback_data="goto_withdraw")]
    ])

    await update.message.reply_text(
        f"{DIV}\n"
        "  💰  Wallet\n"
        f"{DIV}\n\n"
        "💰  AVAILABLE BALANCE\n"
        f"   {fmt_usd(balance)}\n\n"
        f"{DIV_SHORT}\n\n"
        "📊  TASK SUMMARY\n"
        f"   Approved Tasks   {approved}\n"
        f"   In Review        {in_review}\n"
        f"   Total Submitted  {total_submitted}\n\n"
        f"{DIV_SHORT}\n\n"
        "ℹ️  WITHDRAWAL INFO\n"
        "   Minimum: $1.00\n"
        "   Network Fee: $0.025\n\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME


async def handle_support(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        f"{DIV}\n"
        "  📞  Support\n"
        f"{DIV}\n\n"
        "Our team is here to help you.\n\n"
        "Contact us for:\n"
        "  • Account or task issues\n"
        "  • Rejected submission queries\n"
        "  • Withdrawal problems\n"
        "  • Any other questions\n\n"
        f"{DIV_SHORT}\n\n"
        "📌  Admin Contacts:\n"
        "  @axWorker_Admin\n"
        "  @axWorker_Admin\n\n"
        "⏰  Response time: within a few hours\n\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME


async def handle_new_user(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Legacy handler kept for compatibility — redirects to Dashboard."""
    user_id = update.effective_user.id
    get_user(user_id)
    try:
        db.reference("all_users").child(str(user_id)).set({
            "user_id": user_id,
            "username": update.effective_user.username,
            "joined_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        })
    except Exception as e:
        logger.error(f"Failed to add user to all_users: {e}")
    await update.message.reply_text(
        f"{DIV}\n"
        "  ℹ️  Getting Started\n"
        f"{DIV}\n\n"
        "New to the bot? Here are your guides:\n\n"
        "📱  Instagram 2FA Guide\n"
        "  Contact @axWorker_Admin for the guide\n\n"
        "🍪  Facebook Cookie Guide\n"
        "  Contact @axWorker_Admin for the guide\n\n"
        f"{DIV_SHORT}\n\n"
        "⚠️  Note: The Facebook Cookie task\n"
        "is not currently active in this bot.\n\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD,
        disable_web_page_preview=True,
    )
    return HOME


async def handle_leaderboard(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    settings = get_leaderboard_settings()
    if not settings.get("enabled", True):
        await update.message.reply_text(
            f"{DIV}\n"
            "  🏆  Leaderboard\n"
            f"{DIV}\n\n"
            "The leaderboard is currently offline.\n\n"
            "Please check back later.\n\n"
            f"{DIV}"
        )
        return HOME
    mode = settings.get("mode", "auto")
    leaderboard = get_leaderboard_data()
    if not leaderboard:
        await update.message.reply_text(
            f"{DIV}\n"
            "  🏆  Leaderboard\n"
            f"{DIV}\n\n"
            "Leaderboard data is being prepared.\n\n"
            "Please check back shortly.\n\n"
            f"{DIV}"
        )
        return HOME
    text = format_leaderboard_text(leaderboard, mode)
    await update.message.reply_text(text, parse_mode=None)
    return HOME


async def handle_referrals(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    stats = get_referral_stats(user_id)
    bot_username = context.bot.username
    referral_link = f"[t.me](https://t.me/{bot_username}?start={user_id})"
    total_earned = stats.get("total_earned", 0.0)
    task_price = get_task_price()
    bonus_per_task = round(task_price * 0.08, 6)

    await update.message.reply_text(
        f"{DIV}\n"
        "  👥  Invite Friends\n"
        f"{DIV}\n\n"
        "📊  REFERRAL STATS\n"
        f"   Total Invites    {stats['total']}\n"
        f"   New (last 24h)   {stats['new_last_24h']}\n"
        f"   Total Earned     {fmt_usd(total_earned)}\n\n"
        f"{DIV_SHORT}\n\n"
        "💡  HOW IT WORKS\n"
        "   Share your link with friends.\n"
        "   When they complete a task,\n"
        "   you earn 8% of their reward.\n\n"
        f"   Bonus per task:  {fmt_usd(bonus_per_task)}\n\n"
        f"{DIV_SHORT}\n\n"
        "🔗  YOUR REFERRAL LINK\n"
        f"   {referral_link}\n\n"
        f"{DIV}",
        parse_mode=None
    )
    return HOME


async def handle_settings(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        f"{DIV}\n"
        "  ⚙️  Settings\n"
        f"{DIV}\n\n"
        "Manage your preferences\n"
        "and learn about this bot.\n\n"
        f"{DIV}",
        reply_markup=SETTINGS_KEYBOARD,
    )
    return SETTINGS_MENU


async def handle_profile(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    await update.message.reply_text(
        f"{DIV}\n"
        "  👤  Profile\n"
        f"{DIV}\n\n"
        "🪪  YOUR ACCOUNT\n"
        f"   User ID   {user_id}\n\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME


async def handle_about_bot(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        f"{DIV}\n"
        "  ℹ️  About This Bot\n"
        f"{DIV}\n\n"
        "Name       AX Worker\n"
        "Version    2.0.0\n\n"
        f"{DIV_SHORT}\n\n"
        "Description:\n"
        "A professional task-completion platform\n"
        "where users earn real money by creating\n"
        "Instagram and Facebook accounts.\n\n"
        f"{DIV_SHORT}\n\n"
        "Features:\n"
        "  ✅  Task-based earning system\n"
        "  💰  Instant balance tracking\n"
        "  👥  Referral reward program\n"
        "  🏆  Daily leaderboard prizes\n"
        "  📥  USDT & bKash withdrawals\n"
        "  🔐  Secure OTP verification\n\n"
        f"{DIV_SHORT}\n\n"
        "Support: @axWorker_Admin\n\n"
        "Thank you for using our platform! 🙏\n\n"
        f"{DIV}",
        reply_markup=SETTINGS_KEYBOARD,
    )
    return SETTINGS_MENU

# ─────────────────────────────────────────────
# Submission Approval
# ─────────────────────────────────────────────
def approve_submission(user_id: int, submission_id: str, task_price: float):
    try:
        sub_ref = db.reference(f"submissions/{user_id}/{submission_id}")
        submission = sub_ref.get()
        if not submission:
            return False
        user_data = get_user(user_id)
        new_approved = user_data.get("approved", 0) + 1
        new_balance = round(user_data.get("balance", 0.0) + task_price, 4)
        new_in_review = max(0, user_data.get("in_review", 0) - 1)
        update_user(user_id, {
            "approved": new_approved,
            "balance": new_balance,
            "in_review": new_in_review
        })
        referral_data = get_referral_data(user_id)
        referrer_id = referral_data.get("referred_by")
        if referrer_id:
            referral_reward = round(task_price * 0.08, 6)
            if referral_reward > 0:
                referrer_data = get_user(referrer_id)
                new_referrer_balance = round(referrer_data.get("balance", 0.0) + referral_reward, 6)
                update_user(referrer_id, {"balance": new_referrer_balance})
                referrer_ref_data = get_referral_data(referrer_id)
                new_total_earned = round(referrer_ref_data.get("total_earned", 0.0) + referral_reward, 6)
                update_referral_data(referrer_id, {"total_earned": new_total_earned})
                logger.info(f"Referral reward {referral_reward} given to {referrer_id} for user {user_id}'s approval")
        sub_ref.update({"status": "approved"})
        return True
    except Exception as e:
        logger.error(f"approve_submission failed: {e}")
        return False

# ─────────────────────────────────────────────
# TASKS Flow
# ─────────────────────────────────────────────
async def handle_tasks(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    task_settings = get_task_settings()
    fb_settings = get_fb_task_settings()
    task_price = get_task_price()
    task_1h_price = get_task_1h_price()
    fb_price = get_fb_task_price()

    t1_status = "🟢 Available" if task_settings.get("task_6h_enabled", True) else "🔴 Disabled"
    t2_status = "🟢 Available" if task_settings.get("task_1h_enabled", True) else "🔴 Disabled"
    fb_status = "🟢 Available" if fb_settings.get("fb_task_enabled", True) else "🔴 Disabled"

    # Build keyboard with dynamic prices — never hardcoded
    keyboard = []
    if task_settings.get("task_6h_enabled", True):
        keyboard.append([f"📱 Instagram 2FA — {fmt_usd(task_price)}"])
    if task_settings.get("task_1h_enabled", True):
        keyboard.append([f"⭐ Instagram 2FA Premium — {fmt_usd(task_1h_price)}"])
    if fb_settings.get("fb_task_enabled", True):
        keyboard.append([f"🍪 Facebook Cookie — {fmt_usd(fb_price)}"])
    keyboard.append(["🔙 Back"])

    task_menu = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

    await update.message.reply_text(
        f"{DIV}\n"
        "  📋  Available Tasks\n"
        f"{DIV}\n\n"
        f"📱  Instagram 2FA\n"
        f"    Reward: {fmt_usd(task_price)}   Review: ~1 min\n"
        f"    Status: {t1_status}\n\n"
        f"⭐  Instagram 2FA Premium\n"
        f"    Reward: {fmt_usd(task_1h_price)}   Review: ~60 min\n"
        f"    Status: {t2_status}\n\n"
        f"🍪  Facebook Cookie\n"
        f"    Reward: {fmt_usd(fb_price)}   Review: ~1 min\n"
        f"    Status: {fb_status}\n\n"
        f"{DIV_SHORT}\n\n"
        "Select a task below to begin.\n\n"
        f"{DIV}",
        parse_mode=None,
        reply_markup=task_menu,
    )
    return TASK_MENU


async def handle_task_2fa_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    task_settings = get_task_settings()
    if not task_settings.get("task_6h_enabled", True):
        await update.message.reply_text(
            f"{DIV}\n"
            "  🔴  Task Unavailable\n"
            f"{DIV}\n\n"
            "Instagram 2FA task is currently\n"
            "disabled by the administrator.\n\n"
            "Please check back later.\n\n"
            "Questions? Contact support:\n"
            "  @axWorker_Admin\n\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD
        )
        return HOME

    task_price = get_task_price()
    await update.message.reply_text(
        f"{DIV}\n"
        "  📱  Instagram 2FA Task\n"
        f"{DIV}\n\n"
        f"Reward:        {fmt_usd(task_price)}\n"
        f"Review time:   ~1 minute\n\n"
        f"{DIV_SHORT}\n\n"
        "Instructions:\n"
        "  1. Create a new Instagram account\n"
        "     using a real mobile device.\n"
        "  2. Enable Two-Factor Authentication.\n"
        "  3. Submit your 2FA backup key.\n\n"
        "⚠️  Using personal info will cause\n"
        "    immediate rejection.\n\n"
        f"{DIV_SHORT}\n\n"
        "Ready? Tap Start Task below.\n\n"
        f"{DIV}",
        reply_markup=ReplyKeyboardMarkup(
            [["▶️ Start Task"], ["🔙 Cancel"]],
            resize_keyboard=True
        ),
    )
    return TASK_2FA_INFO


async def handle_task_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    username = generate_username()
    context.user_data["task_username"] = username
    await update.message.reply_text(
        f"{DIV}\n"
        "  📱  Account Credentials\n"
        f"{DIV}\n\n"
        "Use these details to create your\n"
        "Instagram account:\n\n"
        f"Username:\n"
        f"<code>{username}</code>\n\n"
        f"Password:\n"
        f"<code>{get_default_password()}</code>\n\n"
        f"{DIV_SHORT}\n\n"
        "After creating the account and\n"
        "enabling 2FA, send your 2FA key below.\n\n"
        f"{DIV}",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup([["🔙 Cancel"]], resize_keyboard=True),
    )
    return TASK_2FA_AWAIT_KEY


async def handle_task_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.pop("task_username", None)
    await update.message.reply_text(
        f"{DIV}\n"
        "  ❌  Task Cancelled\n"
        f"{DIV}\n\n"
        "No problem. You can start a new\n"
        "task anytime from the Tasks menu.\n\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME


async def handle_2fa_key(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = update.message.text.strip()
    user = update.effective_user

    if raw in ("🔙 Cancel", "Cancel ❌"):
        return await handle_task_cancel(update, context)

    # 2FA key validation
    cleaned_key, original_key, error_msg = validate_2fa_key(raw)
    if error_msg:
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Invalid 2FA Key\n"
            f"{DIV}\n\n"
            f"{error_msg}\n\n"
            f"{DIV}",
            reply_markup=ReplyKeyboardMarkup([["🔙 Cancel"]], resize_keyboard=True),
        )
        return TASK_2FA_AWAIT_KEY

    key = cleaned_key
    try:
        otp_code = generate_totp(cleaned_key)
    except Exception:
        otp_code = "123456"

    context.user_data["pending_2fa_key"] = key
    context.user_data["pending_username"] = context.user_data.get("task_username", generate_username())
    context.user_data["pending_otp_code"] = otp_code

    inline_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Copy Code", callback_data=f"copy_otp:{otp_code}")]
    ])
    reply_keyboard = ReplyKeyboardMarkup(
        [["✅ Account Registered"], ["❌ Cancel Task"]],
        resize_keyboard=True,
        one_time_keyboard=True
    )
    await update.message.reply_text(
        f"{DIV}\n"
        "  🔐  Your Verification Code\n"
        f"{DIV}\n\n"
        f"<code>{otp_code}</code>\n\n"
        f"{DIV_SHORT}\n\n"
        "Steps:\n"
        f"  1. Open Instagram\n"
        f"  2. Username: <code>{context.user_data['pending_username']}</code>\n"
        f"  3. Password: <code>{get_default_password()}</code>\n"
        "  4. Enter the code above\n\n"
        "Once logged in, tap Account Registered.\n\n"
        f"{DIV}",
        parse_mode="HTML",
        reply_markup=reply_keyboard
    )
    await update.message.reply_text(
        "👇 Tap to copy your code:",
        reply_markup=inline_keyboard
    )
    return TASK_2FA_STARTED
    
async def handle_2fa_key_1h(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    raw = update.message.text.strip()
    if raw in ("🔙 Cancel", "Cancel ❌"):
        return await handle_task_1h_cancel(update, context)

    # 2FA key validation
    cleaned_key, original_key, error_msg = validate_2fa_key(raw)
    if error_msg:
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Invalid 2FA Key\n"
            f"{DIV}\n\n"
            f"{error_msg}\n\n"
            f"{DIV}",
            reply_markup=ReplyKeyboardMarkup([["🔙 Cancel"]], resize_keyboard=True),
        )
        return TASK_2FA_1H_AWAIT_KEY

    key = cleaned_key
    try:
        otp_code = generate_totp(cleaned_key)
    except Exception:
        otp_code = "123456"

    context.user_data["pending_2fa_key"] = key
    context.user_data["pending_username"] = context.user_data.get("task_username", generate_username())
    context.user_data["pending_otp_code"] = otp_code
    context.user_data["current_task_price"] = get_task_1h_price()

    inline_keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Copy Code", callback_data=f"copy_otp:{otp_code}")]
    ])
    reply_keyboard = ReplyKeyboardMarkup(
        [["✅ Account Registered"], ["❌ Cancel Task"]],
        resize_keyboard=True,
        one_time_keyboard=True
    )
    await update.message.reply_text(
        f"{DIV}\n"
        "  🔐  Your Verification Code\n"
        f"{DIV}\n\n"
        f"<code>{otp_code}</code>\n\n"
        f"{DIV_SHORT}\n\n"
        "Steps:\n"
        f"  1. Open Instagram\n"
        f"  2. Username: <code>{context.user_data['pending_username']}</code>\n"
        f"  3. Password: <code>{get_default_password()}</code>\n"
        "  4. Enter the code above\n\n"
        "Once logged in, tap Account Registered.\n\n"
        f"{DIV}",
        parse_mode="HTML",
        reply_markup=reply_keyboard
    )
    await update.message.reply_text(
        "👇 Tap to copy your code:",
        reply_markup=inline_keyboard
    )
    return TASK_2FA_1H_STARTED


async def _auto_approve_job(bot, user_id: int, sub_id: str, admin_msg_ids: dict) -> None:
    await asyncio.sleep(60)
    sub = db.reference(f"submissions/{user_id}/{sub_id}").get()
    if not sub or sub.get("status", "pending") != "pending":
        return

    task_price = get_task_price()
    success = approve_submission(user_id, sub_id, task_price)
    if not success:
        return

    try:
        await bot.send_message(
            chat_id=user_id,
            text=(
                f"{DIV}\n"
                "  ✅  Submission Approved\n"
                f"{DIV}\n\n"
                f"Your task has been approved!\n\n"
                f"Reward:   +{fmt_usd(task_price)}\n\n"
                "The amount has been added to\n"
                "your wallet. Keep it up! 🎉\n\n"
                f"{DIV}"
            ),
            parse_mode=None,
        )
    except Exception as e:
        logger.error(f"_auto_approve_job notify user failed: {e}")

    for admin_id_str, msg_id in admin_msg_ids.items():
        try:
            await bot.edit_message_text(
                chat_id=int(admin_id_str),
                message_id=msg_id,
                text=(
                    f"✅ AUTO-APPROVED (1 min elapsed)\n\n"
                    f"User ID: {user_id}\n"
                    f"Reward: {fmt_usd(task_price)} paid"
                ),
                parse_mode=None,
            )
        except Exception as e:
            logger.error(f"_auto_approve_job edit admin msg failed: {e}")


async def handle_account_registered(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    key = context.user_data.get("pending_2fa_key")
    ig_username = context.user_data.get("pending_username", generate_username())
    if not key:
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Session Expired\n"
            f"{DIV}\n\n"
            "Your session has timed out.\n"
            "Please start the task again.\n\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD
        )
        return HOME

    password = get_default_password()
    tg_username = f"@{user.username}" if user.username else str(user.id)
    try:
        sub_ref = db.reference(f"submissions/{user.id}").push()
        sub_id = sub_ref.key
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        sub_ref.set({
            "username": ig_username,
            "password": password,
            "key": key,
            "tg_username": tg_username,
            "user_id": str(user.id),
            "datetime": now,
            "status": "pending",
        })
        u = get_user(user.id)
        update_user(user.id, {
            "in_review": u.get("in_review", 0) + 1,
            "total_submitted": u.get("total_submitted", 0) + 1,
        })
        _rebuild_xlsx(user.id)
    except Exception as e:
        logger.error(f"handle_account_registered add_submission failed: {e}")
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Submission Failed\n"
            f"{DIV}\n\n"
            "Could not save your submission.\n"
            "Please try again.\n\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD,
        )
        return HOME

    context.user_data.pop("pending_2fa_key", None)
    context.user_data.pop("pending_username", None)
    context.user_data.pop("task_username", None)
    context.user_data.pop("pending_otp_code", None)

    task_price = get_task_price()
    submission_text = (
        f"🔔 New Submission (auto-approve in 1 min)\n\n"
        f"User ID:  {user.id}\n"
        f"TG:       {tg_username}\n\n"
        f"Instagram Username:\n"
        f"{ig_username}\n\n"
        f"Password:\n"
        f"{password}\n\n"
        f"2FA Key:\n"
        f"{key}\n\n"
        f"Submitted: {now}\n"
        f"Reward: {fmt_usd(task_price)}"
    )
    inline_kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Approve", callback_data=f"quick_approve:{user.id}:{sub_id}"),
            InlineKeyboardButton("❌ Reject", callback_data=f"quick_cancel:{user.id}:{sub_id}"),
        ]
    ])
    admin_msg_ids = {}
    for admin_id in ADMIN_IDS:
        try:
            sent = await context.bot.send_message(
                chat_id=admin_id,
                text=submission_text,
                parse_mode=None,
                reply_markup=inline_kb,
            )
            admin_msg_ids[str(admin_id)] = sent.message_id
        except Exception as e:
            logger.error(f"Failed to send submission to admin {admin_id}: {e}")

    task = asyncio.create_task(
        _auto_approve_job(context.bot, user.id, sub_id, admin_msg_ids)
    )
    _pending_auto_approve[f"{user.id}_{sub_id}"] = task

    await update.message.reply_text(
        f"{DIV}\n"
        "  ✅  Submission Received\n"
        f"{DIV}\n\n"
        "Your account has been registered\n"
        "and sent for review.\n\n"
        f"Review time:   ~1 minute\n"
        f"Reward:        +{fmt_usd(task_price)}\n\n"
        "You'll be notified upon approval.\n\n"
        f"{DIV}",
        parse_mode=None,
        reply_markup=HOME_KEYBOARD
    )
    return HOME


async def handle_2fa_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.pop("pending_2fa_key", None)
    context.user_data.pop("pending_username", None)
    context.user_data.pop("task_username", None)
    context.user_data.pop("pending_otp_code", None)
    await update.message.reply_text(
        f"{DIV}\n"
        "  ❌  Task Cancelled\n"
        f"{DIV}\n\n"
        "You can start a new task anytime\n"
        "from the Tasks menu.\n\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD
    )
    return HOME

# ─────────────────────────────────────────────
# 1H Task Flow
# ─────────────────────────────────────────────
async def handle_task_2fa_1h_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    task_settings = get_task_settings()
    if not task_settings.get("task_1h_enabled", True):
        await update.message.reply_text(
            f"{DIV}\n"
            "  🔴  Task Unavailable\n"
            f"{DIV}\n\n"
            "Instagram 2FA Premium task is\n"
            "currently disabled.\n\n"
            "Please check back later.\n\n"
            "For help, contact: @axWorker_Admin\n\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD
        )
        return HOME

    task_1h_price = get_task_1h_price()
    context.user_data["current_task_price"] = task_1h_price
    context.user_data["current_task_review_time"] = "60 minutes"

    await update.message.reply_text(
        f"{DIV}\n"
        "  ⭐  Instagram 2FA Premium Task\n"
        f"{DIV}\n\n"
        f"Reward:        {fmt_usd(task_1h_price)}\n"
        f"Review time:   ~60 minutes\n\n"
        f"{DIV_SHORT}\n\n"
        "Instructions:\n"
        "  1. Create a new Instagram account\n"
        "     using a real mobile device.\n"
        "  2. Enable Two-Factor Authentication.\n"
        "  3. Submit your 2FA backup key.\n\n"
        "⚠️  Using personal info will cause\n"
        "    immediate rejection.\n\n"
        f"{DIV_SHORT}\n\n"
        "Ready? Tap Start Task below.\n\n"
        f"{DIV}",
        parse_mode=None,
        reply_markup=ReplyKeyboardMarkup(
            [["▶️ Start Task"], ["🔙 Cancel"]],
            resize_keyboard=True
        ),
    )
    return TASK_2FA_1H_INFO


async def handle_task_1h_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.pop("task_username", None)
    context.user_data.pop("current_task_price", None)
    context.user_data.pop("current_task_review_time", None)
    await update.message.reply_text(
        f"{DIV}\n"
        "  ❌  Task Cancelled\n"
        f"{DIV}\n\n"
        "You can start a new task anytime\n"
        "from the Tasks menu.\n\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME
    
async def callback_fb_copy_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query.from_user.id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    pending_subs = context.user_data.get("fbacts_pending_subs", [])
    
    if data.startswith("fbacts_copy_uid:"):
        parts = data.split(":")
        sub_id = parts[2]
        for sub in pending_subs:
            if sub.get("id") == sub_id:
                await query.answer("✅ See below!", show_alert=False)
                await query.message.reply_text(
                    f"Email/Number:\n<code>{sub.get('uid', 'N/A')}</code>",
                    parse_mode="HTML",
                )
                return
    elif data.startswith("fbacts_copy_password:"):
        parts = data.split(":")
        sub_id = parts[2]
        for sub in pending_subs:
            if sub.get("id") == sub_id:
                await query.answer("✅ See below!", show_alert=False)
                await query.message.reply_text(
                    f"Password:\n<code>{sub.get('password', 'N/A')}</code>",
                    parse_mode="HTML",
                )
                return
                
async def callback_fb_acts_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    
    # Navigation
    if data.startswith("fbacts_prev:"):
        current_idx = int(data.split(":")[1])
        new_idx = max(0, current_idx - 1)
        context.user_data["fbacts_current_index"] = new_idx
        await _send_fb_submission_for_review(update, context, new_idx)
        return
    
    if data.startswith("fbacts_next:"):
        current_idx = int(data.split(":")[1])
        pending_subs = context.user_data.get("fbacts_pending_subs", [])
        new_idx = min(len(pending_subs) - 1, current_idx + 1)
        context.user_data["fbacts_current_index"] = new_idx
        await _send_fb_submission_for_review(update, context, new_idx)
        return
    
    if data == "fbacts_exit":
        for key in ("fbacts_pending_subs", "fbacts_target_id", "fbacts_current_index",
                    "fbacts_approved_count", "fbacts_cancelled_count"):
            context.user_data.pop(key, None)
        await query.edit_message_text(
            "FB review session ended.\n\nUse /fbacts {userid} to start a new review."
        )
        await query.answer()
        return
    
    # Approve
    if data.startswith("fbacts_approve:"):
        _, target_id_str, sub_id, index_str = data.split(":")
        target_id = int(target_id_str)
        pending_subs = context.user_data.get("fbacts_pending_subs", [])
        task_price = get_fb_task_price()
        success = approve_fb_submission(target_id, sub_id, task_price)
        if not success:
            await query.answer("❌ Failed to approve.", show_alert=True)
            return
        logger.info(f"Admin {admin_id} approved FB submission {sub_id} for user {target_id}")
        try:
            fb_user_data = get_fb_user(target_id)
            new_balance = fb_user_data.get("balance", 0.0)
            await context.bot.send_message(
                chat_id=target_id,
                text=(
                    f"{DIV}\n"
                    "  ✅  FB Submission Approved\n"
                    f"{DIV}\n\n"
                    f"Reward:  +{fmt_usd(task_price)}\n"
                    f"Balance: {fmt_usd(new_balance)}\n\n"
                    "Thank you! 🎉\n\n"
                    f"{DIV}"
                ),
                parse_mode=None,
            )
        except Exception as e:
            logger.warning(f"Could not notify user {target_id}: {e}")
        approved_count = context.user_data.get("fbacts_approved_count", 0) + 1
        context.user_data["fbacts_approved_count"] = approved_count
        new_pending = [s for s in pending_subs if s.get("id") != sub_id]
        context.user_data["fbacts_pending_subs"] = new_pending
        current_index = context.user_data.get("fbacts_current_index", 0)
        await _send_fb_submission_for_review(update, context, current_index if new_pending else len(new_pending))
        await query.answer("✅ Approved!")
        return
    
    # Reject
    if data.startswith("fbacts_cancel:"):
        _, target_id_str, sub_id, index_str = data.split(":")
        target_id = int(target_id_str)
        pending_subs = context.user_data.get("fbacts_pending_subs", [])
        try:
            db.reference(f"fb_submissions/{target_id}/{sub_id}").delete()
            logger.info(f"Admin {admin_id} rejected FB submission {sub_id} for user {target_id}")
        except Exception as e:
            logger.error(f"Failed to delete FB submission {sub_id}: {e}")
            await query.answer("❌ Failed to reject.", show_alert=True)
            return
        try:
            fb_user_data = get_fb_user(target_id)
            new_in_review = max(0, fb_user_data.get("in_review", 0) - 1)
            update_fb_user(target_id, {"in_review": new_in_review})
            try:
                await context.bot.send_message(
                    chat_id=target_id,
                    text=(
                        f"{DIV}\n"
                        "  ❌  FB Submission Rejected\n"
                        f"{DIV}\n\n"
                        "Your submission did not meet\n"
                        "our requirements.\n\n"
                        "Tips:\n"
                        "  • Use a real mobile device\n"
                        "  • Submit fresh, valid cookies\n"
                        "  • Use the correct UID\n\n"
                        "Contact: @axWorker_Admin\n\n"
                        f"{DIV}"
                    ),
                    parse_mode=None,
                )
            except Exception as e:
                logger.warning(f"Could not notify user {target_id}: {e}")
        except Exception as e:
            logger.error(f"Failed to update FB user {target_id}: {e}")
        cancelled_count = context.user_data.get("fbacts_cancelled_count", 0) + 1
        context.user_data["fbacts_cancelled_count"] = cancelled_count
        new_pending = [s for s in pending_subs if s.get("id") != sub_id]
        context.user_data["fbacts_pending_subs"] = new_pending
        current_index = context.user_data.get("fbacts_current_index", 0)
        await _send_fb_submission_for_review(update, context, current_index if new_pending else len(new_pending))
        await query.answer("❌ Rejected!")
        return


async def handle_task_1h_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    username = generate_username()
    context.user_data["task_username"] = username
    await update.message.reply_text(
        f"{DIV}\n"
        "  ⭐  Account Credentials\n"
        f"{DIV}\n\n"
        "Use these details to create your\n"
        "Instagram account:\n\n"
        f"Username:\n"
        f"<code>{username}</code>\n\n"
        f"Password:\n"
        f"<code>{get_default_password()}</code>\n\n"
        "After creating the account and\n"
        "enabling 2FA, send your 2FA key below.\n\n"
        f"{DIV}",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup([["🔙 Cancel"]], resize_keyboard=True),
    )
    return TASK_2FA_1H_AWAIT_KEY


async def handle_account_registered_1h(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    key = context.user_data.get("pending_2fa_key")
    ig_username = context.user_data.get("pending_username", generate_username())
    task_price = context.user_data.get("current_task_price", 0.220)
    if not key:
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Session Expired\n"
            f"{DIV}\n\n"
            "Please start the task again.\n\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD
        )
        return HOME

    password = get_default_password()
    tg_username = f"@{user.username}" if user.username else str(user.id)
    try:
        add_submission(user.id, tg_username, ig_username, password, key)
    except Exception as e:
        logger.error(f"handle_account_registered_1h add_submission failed: {e}")
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Submission Failed\n"
            f"{DIV}\n\n"
            "Could not save your submission.\n"
            "Please try again.\n\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD,
        )
        return HOME

    context.user_data.pop("pending_2fa_key", None)
    context.user_data.pop("pending_username", None)
    context.user_data.pop("task_username", None)
    context.user_data.pop("pending_otp_code", None)
    context.user_data.pop("current_task_price", None)
    context.user_data.pop("current_task_review_time", None)

    await update.message.reply_text(
        f"{DIV}\n"
        "  ✅  Submission Received\n"
        f"{DIV}\n\n"
        "Your account has been registered\n"
        "and sent for review.\n\n"
        f"Review time:   ~60 minutes\n"
        f"Reward:        +{fmt_usd(task_price)}\n\n"
        "You'll be notified upon approval.\n\n"
        f"{DIV}",
        parse_mode=None,
        reply_markup=HOME_KEYBOARD
    )
    return HOME
        
async def handle_task_fb_info(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    fb_settings = get_fb_task_settings()
    if not fb_settings.get("fb_task_enabled", True):
        await update.message.reply_text(
            f"{DIV}\n"
            "  🔴  Task Unavailable\n"
            f"{DIV}\n\n"
            "Facebook Cookie task is currently\n"
            "disabled by the administrator.\n\n"
            "Please check back later.\n\n"
            "For help, contact: @axWorker_Admin\n\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD,
        )
        return HOME

    price = get_fb_task_price()
    await update.message.reply_text(
        f"{DIV}\n"
        "  🍪  Facebook Cookie Task\n"
        f"{DIV}\n\n"
        f"Reward:        {fmt_usd(price)}\n"
        f"Review time:   ~1 minute\n\n"
        f"{DIV_SHORT}\n\n"
        "Instructions:\n"
        "  1. We provide credentials.\n"
        "  2. Log in on a real mobile device.\n"
        "  3. Send your Account Email/Phone.\n\n"
        "⚠️  Using personal info will cause\n"
        "    immediate rejection.\n\n"
        f"{DIV_SHORT}\n\n"
        "Ready? Tap Start Task below.\n\n"
        f"{DIV}",
        parse_mode=None,
        reply_markup=ReplyKeyboardMarkup(
            [["▶️ Start Task"], ["🔙 Cancel"]],
            resize_keyboard=True,
        ),
    )
    return TASK_FB_INFO

# ─────────────────────────────────────────────
# WITHDRAW Flow (New)
# ─────────────────────────────────────────────
async def handle_withdraw(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    try:
        data = get_user(user_id)
        balance = data.get("balance", 0.0)
    except Exception:
        balance = 0.0

    pending_total = get_pending_withdrawal_total(user_id)

    await update.message.reply_text(
        f"{DIV}\n"
        "  📥  Withdraw\n"
        f"{DIV}\n\n"
        "💰  BALANCE\n"
        f"   Available          {fmt_usd(balance)}\n"
        f"   Pending Withdrawal {fmt_usd(pending_total)}\n\n"
        f"{DIV_SHORT}\n\n"
        "ℹ️  DETAILS\n"
        f"   Minimum Withdrawal   $1.00\n"
        f"   Network Fee          $0.025\n\n"
        "Select a withdrawal method below.\n\n"
        f"{DIV}",
        parse_mode=None,
        reply_markup=WITHDRAW_NEW_KEYBOARD,
    )
    return WITHDRAW_MENU


async def handle_withdraw_new_usdt(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """User chose USDT-BEP20 from the new withdraw menu — ask for wallet address."""
    user_id = update.effective_user.id
    try:
        data = get_user(user_id)
        balance = data.get("balance", 0.0)
    except Exception:
        balance = 0.0

    context.user_data["withdraw_balance"] = balance
    context.user_data["withdraw_method"] = "usdt"

    await update.message.reply_text(
        f"{DIV}\n"
        "  💎  USDT-BEP20 Withdrawal\n"
        f"{DIV}\n\n"
        f"Available:    {fmt_usd(balance)}\n"
        f"Minimum:      $1.00\n"
        f"Network fee:  $0.025\n\n"
        f"{DIV_SHORT}\n\n"
        "Please enter your BEP-20 wallet address:\n\n"
        "⚠️  Double-check before submitting.\n\n"
        f"{DIV}",
        parse_mode=None,
        reply_markup=BACK_KEYBOARD,
    )
    return WITHDRAW_ADDRESS


async def handle_withdraw_new_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """User pressed Cancel ❌ on the new withdraw menu."""
    context.user_data.pop("withdraw_balance", None)
    context.user_data.pop("withdraw_method", None)
    context.user_data.pop("withdraw_amount", None)
    context.user_data.pop("withdraw_wallet", None)
    await update.message.reply_text(
        f"{DIV}\n"
        "  🏠  Main Menu\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME


async def handle_withdraw_bep20(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    """Legacy BEP20 handler from old WITHDRAW_MENU_KEYBOARD (kept for backward compat)."""
    user_id = update.effective_user.id
    try:
        data = get_user(user_id)
        balance = data.get("balance", 0.0)
    except Exception as e:
        logger.error(f"handle_withdraw_bep20 failed: {e}")
        balance = 0.0
    context.user_data["withdraw_balance"] = balance
    context.user_data["withdraw_method"] = "usdt"
    await update.message.reply_text(
        f"{DIV}\n"
        "  💎  USDT Withdrawal\n"
        f"{DIV}\n\n"
        f"Available:   {fmt_usd(balance)}\n"
        f"Minimum:     $1.00\n"
        f"Network fee: $0.025\n\n"
        "Enter the amount you wish to withdraw\n"
        "(in USD):\n\n"
        f"{DIV}",
        reply_markup=BACK_KEYBOARD,
    )
    return WITHDRAW_AMOUNT


async def handle_withdraw_bkash(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user_id = update.effective_user.id
    try:
        data = get_user(user_id)
        balance = data.get("balance", 0.0)
    except Exception as e:
        logger.error(f"handle_withdraw_bkash failed: {e}")
        balance = 0.0
    bdt_balance = balance * USD_TO_BDT_RATE
    context.user_data["withdraw_balance"] = balance
    context.user_data["withdraw_method"] = "bkash"
    await update.message.reply_text(
        f"{DIV}\n"
        "  📱  bKash Withdrawal\n"
        f"{DIV}\n\n"
        f"Balance:     {fmt_usd(balance)} ({fmt_bdt(bdt_balance)})\n"
        f"Minimum:     $1.00 ({fmt_bdt(USD_TO_BDT_RATE)})\n"
        f"Fee:         $0.025 ({fmt_bdt(0.025 * USD_TO_BDT_RATE)})\n"
        f"Rate:        1 USD = {USD_TO_BDT_RATE} BDT\n\n"
        "Enter the amount in USD:\n"
        "Example: 0.50\n\n"
        f"{DIV}",
        parse_mode=None,
        reply_markup=BACK_KEYBOARD,
    )
    return WITHDRAW_AMOUNT


async def handle_withdraw_amount(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "🔙 Back":
        await update.message.reply_text(
            f"{DIV}\n"
            "  🏠  Main Menu\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD
        )
        return HOME
    try:
        amount = float(text)
    except ValueError:
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Invalid Amount\n"
            f"{DIV}\n\n"
            "Please enter a valid number.\n"
            "Example: 1.50\n\n"
            f"{DIV}"
        )
        return WITHDRAW_AMOUNT

    balance = context.user_data.get("withdraw_balance", 0.0)
    if amount < 1.0:
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Minimum Not Met\n"
            f"{DIV}\n\n"
            "Minimum withdrawal amount is $1.00.\n\n"
            f"Your balance: {fmt_usd(balance)}\n\n"
            f"{DIV}"
        )
        return WITHDRAW_AMOUNT
    if amount > balance:
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Insufficient Balance\n"
            f"{DIV}\n\n"
            "You don't have enough balance\n"
            "for this withdrawal.\n\n"
            f"Available: {fmt_usd(balance)}\n\n"
            f"{DIV}"
        )
        return WITHDRAW_AMOUNT

    context.user_data["withdraw_amount"] = amount
    method = context.user_data.get("withdraw_method", "usdt")
    wallet = context.user_data.get("withdraw_wallet", "")
    fee = 0.025
    receive = round(amount - fee, 4)
    user_id = update.effective_user.id

    if method == "usdt" and wallet:
        # New flow: wallet was collected first — show confirmation page
        inline_kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Confirm", callback_data=f"wd_confirm:{user_id}"),
                InlineKeyboardButton("❌ Cancel",  callback_data=f"wd_confirm_cancel:{user_id}"),
            ]
        ])
        short_wallet = f"{wallet[:10]}...{wallet[-6:]}" if len(wallet) > 16 else wallet
        await update.message.reply_text(
            f"{DIV}\n"
            "  📋  Withdrawal Summary\n"
            f"{DIV}\n\n"
            f"User ID:        {user_id}\n"
            f"Method:         USDT (BEP-20)\n"
            f"Wallet:         {short_wallet}\n\n"
            f"{DIV_SHORT}\n\n"
            f"Amount:         {fmt_usd(amount)}\n"
            f"Network fee:    $0.0250\n"
            f"You receive:    {fmt_usd(receive)}\n\n"
            f"Available after: {fmt_usd(balance - amount)}\n\n"
            f"{DIV_SHORT}\n\n"
            "Confirm your withdrawal below.\n\n"
            f"{DIV}",
            parse_mode=None,
            reply_markup=inline_kb,
        )
        return WITHDRAW_CONFIRM

    # Legacy flow (bkash / old USDT) — ask for address next
    if method == "bkash":
        await update.message.reply_text(
            f"{DIV}\n"
            "  📱  bKash Account Number\n"
            f"{DIV}\n\n"
            "Enter your bKash number:\n"
            "Example: 01XXXXXXXXX\n\n"
            "⚠️  Make sure the number is correct.\n\n"
            f"{DIV}",
            reply_markup=BACK_KEYBOARD,
        )
    else:
        await update.message.reply_text(
            f"{DIV}\n"
            "  💎  USDT Wallet Address\n"
            f"{DIV}\n\n"
            "Enter your BEP-20 wallet address:\n\n"
            "⚠️  Double-check before submitting.\n\n"
            f"{DIV}",
            reply_markup=BACK_KEYBOARD,
        )
    return WITHDRAW_ADDRESS


async def handle_withdraw_address(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text == "🔙 Back":
        await update.message.reply_text(
            f"{DIV}\n"
            "  🏠  Main Menu\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD
        )
        return HOME

    method = context.user_data.get("withdraw_method", "usdt")

    # ── New USDT flow: address collected first, amount comes next ─────────────
    if method == "usdt" and "withdraw_amount" not in context.user_data:
        # BEP20 address validation
        if not is_valid_bep20(text):
            await update.message.reply_text(
                f"{DIV}\n"
                "  ⚠️  Invalid Wallet Address\n"
                f"{DIV}\n\n"
                "Please enter a valid BEP-20 wallet address.\n"
                "Format: 0x + 40 hexadecimal characters\n\n"
                "Example: 0x742d35Cc6634C0532925a3b844Bc9e7598f0b0e0\n\n"
                f"{DIV}",
                reply_markup=BACK_KEYBOARD,
            )
            return WITHDRAW_ADDRESS

        context.user_data["withdraw_wallet"] = text
        balance = context.user_data.get("withdraw_balance", 0.0)
        short_wallet = f"{text[:10]}...{text[-6:]}" if len(text) > 16 else text
        await update.message.reply_text(
            f"{DIV}\n"
            "  💎  Withdrawal Amount\n"
            f"{DIV}\n\n"
            f"Wallet:       {short_wallet}\n"
            f"Available:    {fmt_usd(balance)}\n"
            f"Minimum:      $1.00\n"
            f"Network fee:  $0.025\n\n"
            "Enter the amount to withdraw (in USD):\n"
            "Example: 1.50\n\n"
            f"{DIV}",
            parse_mode=None,
            reply_markup=BACK_KEYBOARD,
        )
        return WITHDRAW_AMOUNT

    # ── Legacy flow: amount collected first, address is last step ─────────────
    amount = context.user_data.get("withdraw_amount", 0.0)
    fee = 0.025
    receive_usd = round(amount - fee, 4)

    if method == "bkash":
        bkash_pattern = re.compile(r"^01[3-9]\d{8}$")
        if not bkash_pattern.match(text):
            await update.message.reply_text(
                f"{DIV}\n"
                "  ⚠️  Invalid bKash Number\n"
                f"{DIV}\n\n"
                "Please enter a valid Bangladeshi\n"
                "mobile number.\n\n"
                "Example: 01712345678\n\n"
                f"{DIV}"
            )
            return WITHDRAW_ADDRESS
        wallet = text
        receive_bdt = round(receive_usd * USD_TO_BDT_RATE, 2)
        receive_display = f"{fmt_bdt(receive_bdt)} ({fmt_usd(receive_usd)})"
    else:
        # BEP20 address validation
        if not is_valid_bep20(text):
            await update.message.reply_text(
                f"{DIV}\n"
                "  ⚠️  Invalid Wallet Address\n"
                f"{DIV}\n\n"
                "Please enter a valid BEP-20 wallet address.\n"
                "Format: 0x + 40 hexadecimal characters\n\n"
                "Example: 0x742d35Cc6634C0532925a3b844Bc9e7598f0b0e0\n\n"
                f"{DIV}",
                reply_markup=BACK_KEYBOARD,
            )
            return WITHDRAW_ADDRESS
        wallet = text
        receive_display = fmt_usd(receive_usd)

    user = update.effective_user
    tg_username = f"@{user.username}" if user.username else str(user.id)
    try:
        # Deduct balance immediately
        user_data = get_user(user.id)
        new_balance = max(0.0, round(user_data.get("balance", 0.0) - amount, 4))
        update_user(user.id, {"balance": new_balance})

        w_id = create_withdrawal(user.id, tg_username, amount, wallet)
        db.reference(f"withdrawals/{user.id}/{w_id}").update({"method": method})
    except Exception as e:
        logger.error(f"create_withdrawal (legacy) failed: {e}")
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Request Failed\n"
            f"{DIV}\n\n"
            "Could not submit your withdrawal.\n"
            "Please try again.\n\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD,
        )
        return HOME

    context.user_data.clear()

    if method == "bkash":
        await update.message.reply_text(
            f"{DIV}\n"
            "  ✅  Withdrawal Requested\n"
            f"{DIV}\n\n"
            f"Method:    bKash\n"
            f"Number:    {wallet}\n"
            f"Deducted:  {fmt_usd(amount)}\n"
            f"Fee:       $0.0250\n"
            f"Receive:   {receive_display}\n\n"
            "Processing time: 24–48 hours\n\n"
            f"{DIV}",
            parse_mode=None,
            reply_markup=HOME_KEYBOARD,
        )
        admin_text = (
            f"🔔  Withdrawal Request\n\n"
            f"Withdrawal ID: {w_id}\n"
            f"User ID:  {user.id}\n"
            f"TG:       {tg_username}\n"
            f"Method:   bKash\n"
            f"Number:   {wallet}\n"
            f"Amount:   {fmt_usd(amount)}\n"
            f"Fee:      $0.025\n"
            f"Receive:  {receive_display}\n"
            f"Status:   Pending\n"
            f"Date:     {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        )
    else:
        await update.message.reply_text(
            f"{DIV}\n"
            "  ✅  Withdrawal Requested\n"
            f"{DIV}\n\n"
            f"Method:    USDT (BEP-20)\n"
            f"Wallet:    {wallet[:10]}...{wallet[-6:]}\n"
            f"Deducted:  {fmt_usd(amount)}\n"
            f"Fee:       $0.0250\n"
            f"Receive:   {receive_display}\n\n"
            "Processing time: 24–48 hours\n\n"
            f"{DIV}",
            parse_mode=None,
            reply_markup=HOME_KEYBOARD,
        )
        admin_text = (
            f"🔔  Withdrawal Request\n\n"
            f"Withdrawal ID: {w_id}\n"
            f"User ID:  {user.id}\n"
            f"TG:       {tg_username}\n"
            f"Method:   USDT (BEP-20)\n"
            f"Wallet:   {wallet}\n"
            f"Amount:   {fmt_usd(amount)}\n"
            f"Fee:      $0.025\n"
            f"Receive:  {fmt_usd(receive_usd)}\n"
            f"Status:   Pending\n"
            f"Date:     {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        )

    inline_kb = InlineKeyboardMarkup([
        [
            InlineKeyboardButton("✅ Success", callback_data=f"wd_approve:{user.id}:{w_id}"),
            InlineKeyboardButton("🚫 Cancel",  callback_data=f"wd_cancel:{user.id}:{w_id}"),
        ]
    ])
    for admin_id in ADMIN_IDS:
        try:
            await context.bot.send_message(
                chat_id=admin_id,
                text=admin_text,
                parse_mode=None,
                reply_markup=inline_kb,
            )
        except Exception as e:
            logger.warning(f"Could not notify admin {admin_id}: {e}")
    return HOME


async def callback_withdraw_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handles ✅ Confirm / ❌ Cancel on the withdrawal confirmation inline keyboard."""
    query = update.callback_query
    data = query.data
    user = query.from_user

    if data.startswith("wd_confirm_cancel:"):
        # User cancelled confirmation — clean up and return to home
        context.user_data.pop("withdraw_balance", None)
        context.user_data.pop("withdraw_method", None)
        context.user_data.pop("withdraw_amount", None)
        context.user_data.pop("withdraw_wallet", None)
        await query.answer("Cancelled.")
        await query.edit_message_text(
            f"{DIV}\n"
            "  ❌  Withdrawal Cancelled\n"
            f"{DIV}\n\n"
            "Your withdrawal has been cancelled.\n"
            "No funds were deducted.\n\n"
            f"{DIV}"
        )
        return

    if data.startswith("wd_confirm:"):
        amount = context.user_data.get("withdraw_amount", 0.0)
        wallet = context.user_data.get("withdraw_wallet", "")
        balance = context.user_data.get("withdraw_balance", 0.0)
        fee = 0.025
        receive = round(amount - fee, 4)
        tg_username = f"@{user.username}" if user.username else str(user.id)

        if not wallet or amount <= 0:
            await query.answer("⚠️ Session expired. Please try again.", show_alert=True)
            await query.edit_message_text("Session expired. Please start the withdrawal again.")
            return

        try:
            # Deduct balance immediately
            user_data = get_user(user.id)
            current_balance = user_data.get("balance", 0.0)
            if amount > current_balance:
                await query.answer("⚠️ Insufficient balance.", show_alert=True)
                await query.edit_message_text(
                    f"⚠️ Insufficient balance.\n\n"
                    f"Available: {fmt_usd(current_balance)}\n"
                    f"Requested: {fmt_usd(amount)}"
                )
                return
            new_balance = max(0.0, round(current_balance - amount, 4))
            update_user(user.id, {"balance": new_balance})

            w_id = create_withdrawal(user.id, tg_username, amount, wallet)
            db.reference(f"withdrawals/{user.id}/{w_id}").update({"method": "usdt"})
        except Exception as e:
            logger.error(f"wd_confirm create_withdrawal failed: {e}")
            await query.answer("⚠️ Failed to submit. Please try again.", show_alert=True)
            return

        context.user_data.pop("withdraw_balance", None)
        context.user_data.pop("withdraw_method", None)
        context.user_data.pop("withdraw_amount", None)
        context.user_data.pop("withdraw_wallet", None)

        now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
        short_wallet = f"{wallet[:10]}...{wallet[-6:]}" if len(wallet) > 16 else wallet

        await query.answer("✅ Submitted!")
        await query.edit_message_text(
            f"{DIV}\n"
            "  ✅  Withdrawal Submitted\n"
            f"{DIV}\n\n"
            f"Withdrawal ID:  {w_id}\n"
            f"Method:         USDT (BEP-20)\n"
            f"Wallet:         {short_wallet}\n\n"
            f"Amount:         {fmt_usd(amount)}\n"
            f"Fee:            $0.0250\n"
            f"You receive:    {fmt_usd(receive)}\n\n"
            f"Status:         Pending\n"
            f"Submitted:      {now_str}\n\n"
            "Processing time: 24–48 hours.\n"
            "You will be notified on completion.\n\n"
            f"{DIV}"
        )

        # Notify admins
        admin_text = (
            f"🔔  Withdrawal Request\n\n"
            f"Withdrawal ID: {w_id}\n"
            f"User ID:  {user.id}\n"
            f"TG:       {tg_username}\n"
            f"Method:   USDT (BEP-20)\n"
            f"Wallet:   {wallet}\n"
            f"Amount:   {fmt_usd(amount)}\n"
            f"Fee:      $0.025\n"
            f"Receive:  {fmt_usd(receive)}\n"
            f"Status:   Pending\n"
            f"Date:     {now_str}"
        )
        admin_kb = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✅ Success", callback_data=f"wd_approve:{user.id}:{w_id}"),
                InlineKeyboardButton("🚫 Cancel",  callback_data=f"wd_cancel:{user.id}:{w_id}"),
            ]
        ])
        for admin_id in ADMIN_IDS:
            try:
                await context.bot.send_message(
                    chat_id=admin_id,
                    text=admin_text,
                    parse_mode=None,
                    reply_markup=admin_kb,
                )
            except Exception as e:
                logger.warning(f"Could not notify admin {admin_id}: {e}")

# ─────────────────────────────────────────────
# Back → HOME
# ─────────────────────────────────────────────
async def handle_back_to_home(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await update.message.reply_text(
        f"{DIV}\n"
        "  🏠  Main Menu\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD
    )
    return HOME

# ─────────────────────────────────────────────
# Inline Callbacks — Withdrawal
# ─────────────────────────────────────────────
async def callback_withdrawal(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    await query.answer()
    data = query.data
    try:
        action, user_id_str, w_id = data.split(":")
        user_id = int(user_id_str)
    except Exception:
        await query.edit_message_text("❌ Invalid callback data.")
        return
    try:
        wd = get_withdrawal(user_id, w_id)
    except Exception as e:
        logger.error(f"callback_withdrawal get_withdrawal failed: {e}")
        await query.edit_message_text("❌ Could not fetch withdrawal record.")
        return
    if not wd:
        await query.edit_message_text("❌ Withdrawal record not found.")
        return
    wd_ref = db.reference(f"withdrawals/{user_id}/{w_id}")

    def _atomic_status_update(current_data):
        if current_data is None:
            return None
        if current_data.get("status") != "pending":
            return None
        current_data["status"] = action
        return current_data

    try:
        result = wd_ref.transaction(_atomic_status_update)
    except Exception as e:
        logger.error(f"Firebase transaction failed: {e}")
        await query.edit_message_text("❌ Transaction error. Please try again.")
        return

    if result is None:
        existing = get_withdrawal(user_id, w_id)
        current_status = existing.get("status", "unknown") if existing else "unknown"
        await query.answer(
            f"⚠️ Already processed ({current_status}).",
            show_alert=True,
        )
        await query.edit_message_text(
            query.message.text + f"\n\nAlready {current_status}."
        )
        return

    if action == "wd_cancel":
        # Refund the full amount to the user
        refund_amount = wd.get("amount", 0.0)
        try:
            user_data = get_user(user_id)
            refunded_balance = round(user_data.get("balance", 0.0) + refund_amount, 4)
            update_user(user_id, {"balance": refunded_balance})
        except Exception as e:
            logger.error(f"callback_withdrawal refund failed: {e}")

        update_withdrawal(user_id, w_id, {"status": "cancelled"})
        await query.edit_message_text(
            query.message.text + "\n\n🚫 Cancelled by admin. Amount refunded."
        )
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=(
                    f"{DIV}\n"
                    "  🚫  Withdrawal Rejected\n"
                    f"{DIV}\n\n"
                    f"Withdrawal ID:  {w_id}\n\n"
                    f"Your withdrawal request has been\n"
                    f"rejected by an administrator.\n\n"
                    f"Refund:  +{fmt_usd(refund_amount)}\n"
                    "The full amount has been returned\n"
                    "to your wallet.\n\n"
                    "For questions, contact:\n"
                    "  @axWorker_Admin\n\n"
                    f"{DIV}"
                ),
            )
        except Exception as e:
            logger.warning(f"Could not message user {user_id}: {e}")

    elif action == "wd_approve":
        wd = get_withdrawal(user_id, w_id)
        amount = wd.get("amount", 0.0)
        receive = wd.get("receive", 0.0)
        wallet = wd.get("wallet", "N/A")
        short_wallet = f"{wallet[:10]}...{wallet[-6:]}" if len(wallet) > 16 else wallet
        try:
            update_withdrawal(user_id, w_id, {"status": "approved"})
        except Exception as e:
            logger.error(f"callback_withdrawal approve update failed: {e}")
        tx_hash = generate_tx_hash()
        await query.edit_message_text(
            query.message.text + "\n\n✅ Marked as successful by admin."
        )
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=(
                    f"{DIV}\n"
                    "  ✅  Withdrawal Successful\n"
                    f"{DIV}\n\n"
                    f"Withdrawal ID:  {w_id}\n"
                    f"Method:         USDT (BEP-20)\n"
                    f"Address:        {short_wallet}\n\n"
                    f"Amount sent:    {fmt_usd(receive)}\n"
                    f"Status:         ✅ Success\n\n"
                    f"Transaction:\n"
                    f"{tx_hash}\n\n"
                    f"{DIV}"
                ),
            )
        except Exception as e:
            logger.warning(f"Could not message user {user_id}: {e}")

# ─────────────────────────────────────────────
# 2FA Inline Callback
# ─────────────────────────────────────────────
async def callback_2fa_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data
    if data.startswith("copy_otp:"):
        otp_code = data.split(":")[1]
        await query.answer("Code sent below!", show_alert=False)
        await query.message.reply_text(
            f"📋 Your 6-digit code:\n"
            f"<code>{otp_code}</code>\n\n"
            "Press and hold to copy.",
            parse_mode="HTML"
        )
        return
    if data == "confirm_registered":
        user = update.effective_user
        key = context.user_data.get("pending_2fa_key")
        ig_username = context.user_data.get("pending_username", generate_username())
        if not key:
            await query.answer("❌ Session expired!", show_alert=True)
            await query.edit_message_text("Session expired. Please start the task again.")
            return
        password = get_default_password()
        tg_username = f"@{user.username}" if user.username else str(user.id)
        try:
            add_submission(user.id, tg_username, ig_username, password, key)
        except Exception as e:
            logger.error(f"callback_2fa_handler add_submission failed: {e}")
            await query.answer("❌ Failed to save submission!", show_alert=True)
            return
        context.user_data.pop("pending_2fa_key", None)
        context.user_data.pop("pending_username", None)
        context.user_data.pop("task_username", None)
        context.user_data.pop("pending_otp_code", None)
        await query.answer("✅ Registered!")
        await query.edit_message_text(
            f"✅ Submission received.\n\n"
            f"Review: ~1 min   Reward: +{fmt_usd(get_task_price())}"
        )
        return
    if data == "cancel_2fa_task":
        context.user_data.pop("pending_2fa_key", None)
        context.user_data.pop("pending_username", None)
        context.user_data.pop("task_username", None)
        context.user_data.pop("pending_otp_code", None)
        await query.answer("Cancelled.")
        await query.edit_message_text("Task cancelled.")
        return

# ─────────────────────────────────────────────
# Admin Helpers
# ─────────────────────────────────────────────
def is_admin(update: Update) -> bool:
    return update.effective_user.id in ADMIN_IDS


async def _send_long_message(bot, chat_id: int, text: str, chunk_size: int = 4000):
    lines = text.split("\n")
    chunk = ""
    for line in lines:
        if len(chunk) + len(line) + 1 > chunk_size:
            await bot.send_message(chat_id=chat_id, text=chunk)
            chunk = line + "\n"
        else:
            chunk += line + "\n"
    if chunk.strip():
        await bot.send_message(chat_id=chat_id, text=chunk)

# ─────────────────────────────────────────────
# Admin Commands
# ─────────────────────────────────────────────
async def cmd_botoff(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    reason = " ".join(args) if args else "No reason provided"
    admin_id = update.effective_user.id
    try:
        set_bot_state(False, admin_id)
        await update.message.reply_text(
            f"🔴 BOT OFFLINE\n\n"
            f"Disabled by: {admin_id}\n"
            f"Time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n"
            f"Reason: {reason}\n\n"
            f"Use /boton to re-enable."
        )
    except Exception as e:
        logger.error(f"cmd_botoff failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_boton(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    admin_id = update.effective_user.id
    try:
        set_bot_state(True, admin_id)
        await update.message.reply_text(
            f"🟢 BOT ONLINE\n\n"
            f"Enabled by: {admin_id}\n"
            f"Time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}\n\n"
            f"Bot is fully operational."
        )
    except Exception as e:
        logger.error(f"cmd_boton failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_on2fa6h(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        db.reference("settings/tasks/task_6h_enabled").set("true")
        db.reference("settings/tasks/last_updated").set(
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        )
        await update.message.reply_text(
            "✅ Instagram 2FA task (standard) ENABLED.\n"
            "Reward: $0.030"
        )
    except Exception as e:
        logger.error(f"cmd_on2fa6h failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_off2fa6h(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        db.reference("settings/tasks/task_6h_enabled").set("false")
        db.reference("settings/tasks/last_updated").set(
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        )
        verify = db.reference("settings/tasks").get()
        await update.message.reply_text(
            f"🔴 Instagram 2FA task (standard) DISABLED.\n"
            f"Firebase: {verify}\n"
            f"Use /on2fa6h to re-enable."
        )
    except Exception as e:
        logger.error(f"cmd_off2fa6h failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_on2fa1h(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        db.reference("settings/tasks/task_1h_enabled").set("true")
        db.reference("settings/tasks/last_updated").set(
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        )
        await update.message.reply_text(
            "✅ Instagram 2FA task (premium) ENABLED.\n"
            "Reward: $0.220"
        )
    except Exception as e:
        logger.error(f"cmd_on2fa1h failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_off2fa1h(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        db.reference("settings/tasks/task_1h_enabled").set("false")
        db.reference("settings/tasks/last_updated").set(
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        )
        verify = db.reference("settings/tasks").get()
        await update.message.reply_text(
            f"🔴 Instagram 2FA task (premium) DISABLED.\n"
            f"Firebase: {verify}\n"
            f"Use /on2fa1h to re-enable."
        )
    except Exception as e:
        logger.error(f"cmd_off2fa1h failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_ldoff(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        set_leaderboard_settings({
            "enabled": False,
            "last_update": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        })
        await update.message.reply_text(
            "🔴 Leaderboard OFF.\n\n"
            "Users will see: 'Leaderboard is offline'.\n"
            "Use /ldauto or /ldset to re-enable."
        )
    except Exception as e:
        logger.error(f"cmd_ldoff failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_ldset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("Generating real-time leaderboard...")
    try:
        leaderboard = generate_real_leaderboard()
        set_leaderboard_data(leaderboard)
        set_leaderboard_settings({
            "mode": "real",
            "last_update": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "enabled": True
        })
        text = format_leaderboard_text(leaderboard, "real")
        await status_msg.edit_text(text, parse_mode=None)
        await update.message.reply_text(
            f"✅ Real-time leaderboard updated.\n"
            f"Time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
        )
    except Exception as e:
        logger.error(f"cmd_ldset failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def cmd_ldauto(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("Generating auto leaderboard...")
    previous = get_leaderboard_data()
    try:
        settings = get_leaderboard_settings()
        is_auto_mode = settings.get("mode") == "auto"
        if is_auto_mode and previous:
            leaderboard = generate_auto_leaderboard(increment=1, previous_leaderboard=previous)
            await update.message.reply_text(
                f"✅ Leaderboard incremented (+3-5 tasks per user).\n"
                f"Time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
            )
        else:
            leaderboard = generate_auto_leaderboard(increment=0, previous_leaderboard=None)
            await update.message.reply_text(
                f"✅ Auto leaderboard generated (10 users).\n"
                f"Time: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
            )
        set_leaderboard_data(leaderboard)
        set_leaderboard_settings({
            "mode": "auto",
            "last_update": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "enabled": True
        })
        text = format_leaderboard_text(leaderboard, "auto")
        await status_msg.edit_text(text, parse_mode=None)
    except Exception as e:
        logger.error(f"cmd_ldauto failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("⏳ Loading statistics, please wait…")
    try:
        bot_state = get_bot_state()
        bot_status = "🟢 ONLINE" if bot_state.get("enabled", True) else "🔴 OFFLINE"
        all_users = get_all_users()
        total_users = len(all_users)
        all_subs = get_all_submissions()
        total_submissions = 0
        pending_review = 0
        for uid, subs in all_subs.items():
            if subs:
                sub_count = len(subs)
                total_submissions += sub_count
                pending_review += sub_count
        total_approved = 0
        total_balance = 0.0
        total_withdrawn = 0.0
        for uid, udata in all_users.items():
            total_approved += udata.get('approved', 0)
            total_balance += udata.get('balance', 0.0)
        try:
            all_withdrawals = db.reference("withdrawals").get()
            if all_withdrawals:
                for uid, wds in all_withdrawals.items():
                    if wds:
                        for w_id, w_data in wds.items():
                            if w_data.get('status') == 'approved':
                                total_withdrawn += w_data.get('amount', 0.0)
        except Exception as e:
            logger.warning(f"Could not fetch withdrawals for stats: {e}")
        task_price = get_task_price()
        completion_rate = (total_approved / total_submissions * 100) if total_submissions > 0 else 0
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        today_submissions = 0
        today_withdrawals = 0
        for uid, subs in all_subs.items():
            if subs:
                for sub_id, sub_data in subs.items():
                    if sub_data.get('datetime', '').startswith(today):
                        today_submissions += 1
        try:
            all_withdrawals = db.reference("withdrawals").get()
            if all_withdrawals:
                for uid, wds in all_withdrawals.items():
                    if wds:
                        for w_id, w_data in wds.items():
                            if w_data.get('datetime', '').startswith(today):
                                today_withdrawals += 1
        except Exception:
            pass
        stats_text = (
            f"BOT STATISTICS\n"
            f"{DIV}\n\n"
            f"Status: {bot_status}\n"
            f"Last change: {bot_state.get('last_updated', 'Never')}\n\n"
            f"USERS\n"
            f"Total: {total_users}\n"
            f"Active: {len([u for u in all_users.values() if u.get('total_submitted', 0) > 0])}\n\n"
            f"SUBMISSIONS\n"
            f"Total: {total_submissions}\n"
            f"Today: {today_submissions}\n"
            f"Pending: {pending_review}\n"
            f"Approved: {total_approved}\n\n"
            f"FINANCIALS\n"
            f"Task price: {fmt_usd(task_price)}\n"
            f"Total balance: {fmt_usd(total_balance)}\n"
            f"Total withdrawn: {fmt_usd(total_withdrawn)}\n"
            f"Pending payout: {fmt_usd(pending_review * task_price)}\n"
            f"Completion rate: {completion_rate:.1f}%\n\n"
            f"PERFORMANCE\n"
            f"Avg/user: {total_submissions/total_users if total_users > 0 else 0:.1f} tasks\n"
            f"Today WD requests: {today_withdrawals}"
        )
        await status_msg.edit_text(stats_text, parse_mode=None)
    except Exception as e:
        logger.error(f"cmd_stats failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def cmd_userinfo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if not context.args:
        await update.message.reply_text("Usage: /userinfo {userid}")
        return
    try:
        user_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID.")
        return
    status_msg = await update.message.reply_text("⏳ Fetching user info…")
    try:
        data = get_user(user_id)
        submissions = get_submissions(user_id)
        submission_count = len(submissions)
        pending_wd = 0
        total_withdrawn = 0.0
        try:
            withdrawals = db.reference(f"withdrawals/{user_id}").get()
            if withdrawals:
                for w_id, w_data in withdrawals.items():
                    status = w_data.get("status", "")
                    amount = w_data.get("amount", 0.0)
                    if status == "pending":
                        pending_wd += 1
                    elif status == "approved":
                        total_withdrawn += amount
        except Exception as e:
            logger.warning(f"Could not fetch withdrawals for {user_id}: {e}")
        tg_username = "N/A"
        if submissions:
            for sub in submissions:
                if sub.get("tg_username"):
                    tg_username = sub.get("tg_username")
                    break
        total_earned = data.get('approved', 0) * get_task_price()
        info_text = (
            f"USER INFO\n"
            f"{DIV}\n\n"
            f"ID:       {user_id}\n"
            f"TG:       {tg_username}\n\n"
            f"Balance:  {fmt_usd(data.get('balance', 0.0))}\n"
            f"Withdrawn:{fmt_usd(total_withdrawn)}\n"
            f"Earned:   {fmt_usd(total_earned)}\n\n"
            f"Approved:   {data.get('approved', 0)}\n"
            f"In review:  {data.get('in_review', 0)}\n"
            f"Submitted:  {data.get('total_submitted', 0)}\n"
            f"Pending:    {submission_count}\n"
            f"Pending WD: {pending_wd}\n\n"
            f"Task price: {fmt_usd(get_task_price())}"
        )
        await status_msg.edit_text(info_text, parse_mode=None)
    except Exception as e:
        logger.error(f"cmd_userinfo failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def cmd_stp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if not context.args:
        current_price = get_task_price()
        await update.message.reply_text(
            f"Current task price: {fmt_usd(current_price)}\n\n"
            f"Usage: /stp 0.030"
        )
        return
    try:
        new_price = float(context.args[0])
        if new_price <= 0:
            await update.message.reply_text("❌ Price must be > 0.")
            return
        old_price = get_task_price()
        set_task_price(new_price)
        await update.message.reply_text(
            f"✅ Task price updated.\n\n"
            f"Old: {fmt_usd(old_price)}\n"
            f"New: {fmt_usd(new_price)}"
        )
    except ValueError:
        await update.message.reply_text("❌ Invalid price. Example: /stp 0.030")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_msg(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if len(context.args) < 2:
        await update.message.reply_text("Usage: /msg userid message")
        return
    try:
        user_id = int(context.args[0])
        message = " ".join(context.args[1:])
        await context.bot.send_message(chat_id=user_id, text=message)
        await update.message.reply_text(f"✅ Sent to {user_id}.")
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID.")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_cast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if len(context.args) < 1:
        await update.message.reply_text("Usage: /cast message")
        return
    message = " ".join(context.args)
    status_msg = await update.message.reply_text("Broadcasting...")
    try:
        all_users = get_all_users()
        if not all_users:
            await status_msg.edit_text("❌ No users found.")
            return
        success_count = 0
        fail_count = 0
        for user_id_str in all_users.keys():
            try:
                user_id = int(user_id_str)
                await context.bot.send_message(chat_id=user_id, text=message)
                success_count += 1
            except Exception as e:
                fail_count += 1
                logger.warning(f"Failed to send broadcast to {user_id_str}: {e}")
        await status_msg.edit_text(
            f"✅ Broadcast done.\n\n"
            f"Sent: {success_count}\n"
            f"Failed: {fail_count}"
        )
    except Exception as e:
        await status_msg.edit_text(f"❌ Error: {e}")


async def cmd_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if len(args) < 2:
        await update.message.reply_text("Usage: /add {amount} {userid}")
        return
    try:
        amount = float(args[0])
        target_id = int(args[1])
    except ValueError:
        await update.message.reply_text("❌ Invalid arguments.")
        return
    try:
        user_data = get_user(target_id)
        new_balance = round(user_data.get("balance", 0.0) + amount, 4)
        update_user(target_id, {"balance": new_balance})
        await update.message.reply_text(
            f"✅ Added {fmt_usd(amount)} to {target_id}.\n"
            f"New balance: {fmt_usd(new_balance)}"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_rm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if len(args) < 2:
        await update.message.reply_text("Usage: /rm {amount} {userid}")
        return
    try:
        amount = float(args[0])
        target_id = int(args[1])
    except ValueError:
        await update.message.reply_text("❌ Invalid arguments.")
        return
    try:
        user_data = get_user(target_id)
        new_balance = max(0.0, round(user_data.get("balance", 0.0) - amount, 4))
        update_user(target_id, {"balance": new_balance})
        await update.message.reply_text(
            f"✅ Removed {fmt_usd(amount)} from {target_id}.\n"
            f"New balance: {fmt_usd(new_balance)}"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_rmreview(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if not args:
        await update.message.reply_text("Usage: /rmreview {userid}")
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID.")
        return
    try:
        remove_submissions(target_id)
        update_user(target_id, {"in_review": 0})
        await update.message.reply_text(f"✅ Submissions cleared for {target_id}.")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_apr(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if len(args) < 2:
        await update.message.reply_text("Usage: /apr {amount} {userid}")
        return
    try:
        amount = int(args[0])
        target_id = int(args[1])
    except ValueError:
        await update.message.reply_text("❌ Invalid arguments.")
        return
    try:
        user_data = get_user(target_id)
        new_approved = user_data.get("approved", 0) + amount
        new_review = max(0, user_data.get("in_review", 0) - amount)
        update_user(target_id, {"approved": new_approved, "in_review": new_review})
        await update.message.reply_text(
            f"✅ +{amount} approved for {target_id}.\n"
            f"Total approved: {new_approved}"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_rcv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if not args:
        await update.message.reply_text("Usage: /rcv {userid}")
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID.")
        return
    subs = get_submissions(target_id)
    if not subs:
        await update.message.reply_text(f"No submissions for user {target_id}.")
        return
    tmp_path = None
    try:
        xlsx_data = build_xlsx_bytes(target_id)
        if not xlsx_data:
            await update.message.reply_text("❌ Could not generate XLSX.")
            return
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(xlsx_data)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=f"submissions_{target_id}.xlsx",
                caption=f"Submissions for {target_id} ({len(subs)} records)",
            )
    except Exception as e:
        logger.error(f"cmd_rcv failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


async def cmd_rcvall(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("⏳ Generating approved submissions XLSX...")
    tmp_path = None
    try:
        all_subs = get_all_submissions()
        if not all_subs:
            await status_msg.edit_text("No submissions found.")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Approved Submissions"
        ws.append(["#", "User ID", "TG Username", "Instagram Username", "Password", "2FA Key", "Date (UTC)", "Status"])
        row_num = 1
        for user_id, submissions in all_subs.items():
            if not isinstance(submissions, dict):
                continue
            for sub_id, sub in submissions.items():
                if not isinstance(sub, dict):
                    continue
                if sub.get("status", "pending") != "approved":
                    continue
                ws.append([
                    row_num,
                    sub.get("user_id", user_id),
                    sub.get("tg_username", ""),
                    sub.get("username", ""),
                    sub.get("password", ""),
                    sub.get("key", ""),
                    sub.get("datetime", ""),
                    "approved",
                ])
                row_num += 1
        if row_num == 1:
            await status_msg.edit_text("No approved submissions found.")
            return
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename="approved_submissions.xlsx",
                caption=f"✅ Total approved: {row_num - 1}",
            )
        await status_msg.delete()
    except Exception as e:
        logger.error(f"cmd_rcvall failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


async def cmd_resetsub(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    try:
        all_subs = get_all_submissions()
        total = sum(
            len(subs) if isinstance(subs, dict) else 0
            for subs in all_subs.values()
        ) if all_subs else 0

        db.reference("submissions").delete()
        try:
            db.reference("xlsx_cache").delete()
        except Exception:
            pass

        all_users = db.reference("users").get() or {}
        for uid in all_users:
            try:
                db.reference(f"users/{uid}").update({"in_review": 0})
            except Exception:
                pass

        await update.message.reply_text(
            f"🗑️ All submissions deleted.\n\n"
            f"Records removed: {total}",
        )
        logger.info(f"Admin {update.effective_user.id} reset all submissions ({total} records).")
    except Exception as e:
        logger.error(f"cmd_resetsub failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_live(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("⏳ Fetching active users…")
    try:
        all_users = get_all_users()
        all_subs = get_all_submissions()
        if not all_users:
            await status_msg.edit_text("No users found.")
            return
        active_users = []
        for uid, udata in all_users.items():
            sub_count = len(all_subs.get(uid, {}))
            if sub_count > 0:
                active_users.append({
                    "user_id": int(uid),
                    "submissions": sub_count,
                    "balance": udata.get('balance', 0.0)
                })
        active_users.sort(key=lambda x: x['submissions'], reverse=True)
        if not active_users:
            await status_msg.edit_text("No active users found.")
            return
        total_users = len(active_users)
        PAGE_SIZE = 10
        total_pages = (total_users + PAGE_SIZE - 1) // PAGE_SIZE
        context.user_data["live_users"] = active_users
        context.user_data["live_total_pages"] = total_pages
        await _send_live_page(update, context, 1)
    except Exception as e:
        logger.error(f"cmd_live failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def _send_live_page(update: Update, context: ContextTypes.DEFAULT_TYPE, page: int):
    active_users = context.user_data.get("live_users", [])
    total_pages = context.user_data.get("live_total_pages", 1)
    PAGE_SIZE = 10
    if not active_users:
        await update.message.reply_text("No active users found.")
        return
    start_idx = (page - 1) * PAGE_SIZE
    end_idx = start_idx + PAGE_SIZE
    page_users = active_users[start_idx:end_idx]
    lines = [
        f"ACTIVE USERS   Page {page}/{total_pages}",
        DIV_SHORT,
        "",
    ]
    for idx, user in enumerate(page_users, start=start_idx + 1):
        lines.append(
            f"{idx}. ID: {user['user_id']}\n"
            f"   Submissions: {user['submissions']}   "
            f"Balance: {fmt_usd(user['balance'])}\n"
        )
    lines.append(DIV_SHORT)
    message_text = "\n".join(lines)
    keyboard = []
    for user in page_users:
        keyboard.append([
            InlineKeyboardButton(
                f"📋 {user['user_id']}",
                callback_data=f"live_copy_uid:{user['user_id']}"
            )
        ])
    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("◀ Prev", callback_data=f"live_page:{page - 1}"))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton("Next ▶", callback_data=f"live_page:{page + 1}"))
    if nav_buttons:
        keyboard.append(nav_buttons)
    keyboard.append([InlineKeyboardButton("✖ Close", callback_data="live_close")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(message_text, parse_mode=None, reply_markup=reply_markup)
        await update.callback_query.answer()
    else:
        await update.message.reply_text(message_text, parse_mode=None, reply_markup=reply_markup)


async def callback_live_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    if data.startswith("live_page:"):
        page = int(data.split(":")[1])
        await _send_live_page(update, context, page)
        return
    if data == "live_close":
        await query.edit_message_text("List closed.")
        await query.answer()
        return
    if data.startswith("live_copy_uid:"):
        uid = data.split(":")[1]
        await query.answer("✅ See below!", show_alert=False)
        await query.message.reply_text(
            f"User ID:\n<code>{uid}</code>\n\n"
            "Press and hold to copy.",
            parse_mode="HTML"
        )
        return


async def cmd_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("Fetching user list...")
    try:
        all_users = get_all_users()
        if not all_users:
            await status_msg.edit_text("No users found.")
            return
        user_ids = sorted([int(uid) for uid in all_users.keys()])
        total_users = len(user_ids)
        PAGE_SIZE = 10
        total_pages = (total_users + PAGE_SIZE - 1) // PAGE_SIZE
        context.user_data["user_list"] = user_ids
        context.user_data["total_pages"] = total_pages
        await _send_user_list_page(update, context, 1)
    except Exception as e:
        logger.error(f"cmd_list failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def _send_user_list_page(update: Update, context: ContextTypes.DEFAULT_TYPE, page: int):
    user_ids = context.user_data.get("user_list", [])
    total_pages = context.user_data.get("total_pages", 1)
    PAGE_SIZE = 10
    if not user_ids:
        await update.message.reply_text("No user list. Use /list again.")
        return
    start_idx = (page - 1) * PAGE_SIZE
    end_idx = start_idx + PAGE_SIZE
    page_users = user_ids[start_idx:end_idx]
    lines = [
        f"USER LIST   Total: {len(user_ids)}   Page {page}/{total_pages}",
        DIV_SHORT,
        "",
    ]
    for idx, uid in enumerate(page_users, start=start_idx + 1):
        lines.append(f"{idx}. `{uid}`")
    lines += ["", "Tap a button to copy the user ID."]
    message_text = "\n".join(lines)
    keyboard = []
    row = []
    for uid in page_users:
        row.append(InlineKeyboardButton(str(uid), callback_data=f"copy_uid:{uid}"))
        if len(row) == 3:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)
    nav_buttons = []
    if page > 1:
        nav_buttons.append(InlineKeyboardButton("◀ Prev", callback_data=f"list_page:{page - 1}"))
    if page < total_pages:
        nav_buttons.append(InlineKeyboardButton("Next ▶", callback_data=f"list_page:{page + 1}"))
    if nav_buttons:
        keyboard.append(nav_buttons)
    keyboard.append([InlineKeyboardButton("✖ Close", callback_data="list_close")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(
            message_text, parse_mode=None, reply_markup=reply_markup
        )
        await update.callback_query.answer()
    else:
        await update.message.reply_text(
            message_text, parse_mode=None, reply_markup=reply_markup
        )


async def callback_list_handlers(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = query.from_user.id
    if user_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    if data.startswith("list_page:"):
        page = int(data.split(":")[1])
        await _send_user_list_page(update, context, page)
        return
    if data == "list_close":
        await query.edit_message_text("List closed.")
        await query.answer()
        return
    if data.startswith("copy_uid:"):
        uid = data.split(":")[1]
        await query.answer(f"✅ See below!", show_alert=False)
        await query.message.reply_text(
            f"User ID:\n`{uid}`\n\n"
            f"Use with:\n"
            f"`/add 0.5 {uid}`\n"
            f"`/userinfo {uid}`\n"
            f"`/msg {uid} Hello`",
            parse_mode=None
        )
        return


async def cmd_acts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    args = context.args
    if not args:
        await update.message.reply_text("Usage: /acts {userid}")
        return
    try:
        target_id = int(args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID.")
        return
    context.user_data["acts_target_id"] = target_id
    submissions = get_submissions(target_id)
    if not submissions:
        await update.message.reply_text(f"No pending submissions for user {target_id}.")
        return
    context.user_data["acts_pending_subs"] = submissions
    context.user_data["acts_current_index"] = 0
    context.user_data["acts_approved_count"] = 0
    context.user_data["acts_cancelled_count"] = 0
    await _send_submission_for_review(update, context, 0)
    return ADMIN_ACTS_VIEW


async def _send_submission_for_review(update: Update, context: ContextTypes.DEFAULT_TYPE, index: int):
    pending_subs = context.user_data.get("acts_pending_subs", [])
    target_id = context.user_data.get("acts_target_id")
    if not pending_subs or index >= len(pending_subs):
        approved_count = context.user_data.get("acts_approved_count", 0)
        cancelled_count = context.user_data.get("acts_cancelled_count", 0)
        summary = (
            f"Review Complete\n\n"
            f"User:      {target_id}\n"
            f"Approved:  {approved_count}\n"
            f"Rejected:  {cancelled_count}"
        )
        for key in ("acts_pending_subs", "acts_target_id", "acts_current_index",
                    "acts_approved_count", "acts_cancelled_count"):
            context.user_data.pop(key, None)
        if update.callback_query:
            await update.callback_query.edit_message_text(summary, parse_mode=None)
            await update.callback_query.answer()
        else:
            await update.message.reply_text(summary, parse_mode=None)
        return

    sub = pending_subs[index]
    total = len(pending_subs)
    current = index + 1
    ig_username = sub.get('username', 'N/A')
    password = sub.get('password', 'N/A')
    twofa_key = sub.get('key', 'N/A')
    submission_text = (
        f"SUBMISSION REVIEW   {current}/{total}\n"
        f"{DIV_SHORT}\n"
        f"User: {target_id}\n\n"
        f"Instagram Username:\n"
        f"<code>{ig_username}</code>\n\n"
        f"Password:\n"
        f"<code>{password}</code>\n\n"
        f"2FA Key:\n"
        f"<code>{twofa_key}</code>\n\n"
        f"TG: {sub.get('tg_username', 'N/A')}\n"
        f"At: {sub.get('datetime', 'N/A')}"
    )
    keyboard = [
        [InlineKeyboardButton("📋 Copy Username", callback_data=f"acts_copy_username:{target_id}:{sub.get('id')}:{index}")],
        [InlineKeyboardButton("🔑 Copy Password", callback_data=f"acts_copy_password:{target_id}:{sub.get('id')}:{index}")],
        [InlineKeyboardButton("🔢 Copy 2FA Key", callback_data=f"acts_copy_2fa:{target_id}:{sub.get('id')}:{index}")],
        [
            InlineKeyboardButton("✅ Approve", callback_data=f"acts_approve:{target_id}:{sub.get('id')}:{index}"),
            InlineKeyboardButton("❌ Reject", callback_data=f"acts_cancel:{target_id}:{sub.get('id')}:{index}"),
        ],
        [InlineKeyboardButton("✖ Exit Review", callback_data="acts_exit")],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(submission_text, parse_mode="HTML", reply_markup=reply_markup)
        await update.callback_query.answer()
    else:
        await update.message.reply_text(submission_text, parse_mode="HTML", reply_markup=reply_markup)


async def callback_acts_copy_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    pending_subs = context.user_data.get("acts_pending_subs", [])
    if data.startswith("acts_copy_username:"):
        parts = data.split(":")
        sub_id = parts[2]
        for sub in pending_subs:
            if sub.get('id') == sub_id:
                await query.answer("✅ See below!", show_alert=False)
                await query.message.reply_text(
                    f"Username:\n<code>{sub.get('username', 'N/A')}</code>",
                    parse_mode="HTML"
                )
                return
    elif data.startswith("acts_copy_password:"):
        parts = data.split(":")
        sub_id = parts[2]
        for sub in pending_subs:
            if sub.get('id') == sub_id:
                await query.answer("✅ See below!", show_alert=False)
                await query.message.reply_text(
                    f"Password:\n<code>{sub.get('password', 'N/A')}</code>",
                    parse_mode="HTML"
                )
                return
    elif data.startswith("acts_copy_2fa:"):
        parts = data.split(":")
        sub_id = parts[2]
        for sub in pending_subs:
            if sub.get('id') == sub_id:
                await query.answer("✅ See below!", show_alert=False)
                await query.message.reply_text(
                    f"2FA Key:\n<code>{sub.get('key', 'N/A')}</code>",
                    parse_mode="HTML"
                )
                return


async def callback_quick_review(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return

    data = query.data
    parts = data.split(":")
    if len(parts) != 3:
        await query.answer("❌ Invalid data.", show_alert=True)
        return

    action, user_id_str, sub_id = parts[0], parts[1], parts[2]
    try:
        user_id = int(user_id_str)
    except ValueError:
        await query.answer("❌ Invalid user ID.", show_alert=True)
        return

    task_key = f"{user_id}_{sub_id}"
    pending_task = _pending_auto_approve.pop(task_key, None)
    if pending_task and not pending_task.done():
        pending_task.cancel()

    task_price = get_task_price()

    if action == "quick_approve":
        sub = db.reference(f"submissions/{user_id}/{sub_id}").get()
        if not sub:
            await query.answer("⚠️ Already processed.", show_alert=True)
            await query.edit_message_reply_markup(reply_markup=None)
            return

        success = approve_submission(user_id, sub_id, task_price)
        if success:
            await query.edit_message_text(
                f"✅ APPROVED by admin\n\n"
                f"User: {user_id}\n"
                f"Reward: {fmt_usd(task_price)} paid",
                parse_mode=None,
            )
            await query.answer("✅ Approved!")
            try:
                await context.bot.send_message(
                    chat_id=user_id,
                    text=(
                        f"{DIV}\n"
                        "  ✅  Submission Approved\n"
                        f"{DIV}\n\n"
                        f"Reward: +{fmt_usd(task_price)}\n\n"
                        "The amount has been added to\n"
                        "your wallet. Keep it up! 🎉\n\n"
                        f"{DIV}"
                    ),
                    parse_mode=None,
                )
            except Exception as e:
                logger.error(f"quick_approve notify user failed: {e}")
        else:
            await query.answer("⚠️ Already processed.", show_alert=True)

    elif action == "quick_cancel":
        sub = db.reference(f"submissions/{user_id}/{sub_id}").get()
        if not sub:
            await query.answer("⚠️ Already processed.", show_alert=True)
            await query.edit_message_reply_markup(reply_markup=None)
            return

        try:
            db.reference(f"submissions/{user_id}/{sub_id}").update({"status": "rejected"})
            u = get_user(user_id)
            new_review = max(0, u.get("in_review", 0) - 1)
            update_user(user_id, {"in_review": new_review})
        except Exception as e:
            logger.error(f"quick_cancel update submission failed: {e}")

        await query.edit_message_text(
            f"❌ REJECTED by admin\n\nUser: {user_id}",
            parse_mode=None,
        )
        await query.answer("❌ Rejected!")
        try:
            await context.bot.send_message(
                chat_id=user_id,
                text=(
                    f"{DIV}\n"
                    "  ❌  Submission Rejected\n"
                    f"{DIV}\n\n"
                    "Your submission did not meet\n"
                    "our requirements.\n\n"
                    "Please try again with a valid\n"
                    "account. Contact support if\n"
                    "you need help.\n\n"
                    f"{DIV}"
                ),
                parse_mode=None,
            )
        except Exception as e:
            logger.error(f"quick_cancel notify user failed: {e}")


async def callback_acts_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    admin_id = query.from_user.id
    if admin_id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    if data == "acts_exit":
        for key in ("acts_pending_subs", "acts_target_id", "acts_current_index",
                    "acts_approved_count", "acts_cancelled_count"):
            context.user_data.pop(key, None)
        await query.edit_message_text(
            "Review session ended.\n\nUse /acts {userid} to start a new review."
        )
        await query.answer()
        return
    if data.startswith("acts_approve:"):
        _, target_id_str, sub_id, index_str = data.split(":")
        target_id = int(target_id_str)
        pending_subs = context.user_data.get("acts_pending_subs", [])
        task_price = get_task_price()
        success = approve_submission(target_id, sub_id, task_price)
        if not success:
            await query.answer("❌ Failed to approve.", show_alert=True)
            return
        logger.info(f"Admin {admin_id} approved submission {sub_id} for user {target_id}")
        try:
            user_data = get_user(target_id)
            new_balance = user_data.get("balance", 0.0)
            await context.bot.send_message(
                chat_id=target_id,
                text=(
                    f"{DIV}\n"
                    "  ✅  Submission Approved\n"
                    f"{DIV}\n\n"
                    f"Reward:  +{fmt_usd(task_price)}\n"
                    f"Balance: {fmt_usd(new_balance)}\n\n"
                    "Thank you for your work! 🎉\n\n"
                    f"{DIV}"
                ),
                parse_mode=None
            )
        except Exception as e:
            logger.warning(f"Could not notify user {target_id}: {e}")
        approved_count = context.user_data.get("acts_approved_count", 0) + 1
        context.user_data["acts_approved_count"] = approved_count
        new_pending = [sub for sub in pending_subs if sub.get('id') != sub_id]
        context.user_data["acts_pending_subs"] = new_pending
        current_index = context.user_data.get("acts_current_index", 0)
        await _send_submission_for_review(update, context, current_index if new_pending else len(new_pending))
        await query.answer("✅ Approved!")
        return
    if data.startswith("acts_cancel:"):
        _, target_id_str, sub_id, index_str = data.split(":")
        target_id = int(target_id_str)
        pending_subs = context.user_data.get("acts_pending_subs", [])
        try:
            db.reference(f"submissions/{target_id}/{sub_id}").delete()
            logger.info(f"Admin {admin_id} rejected submission {sub_id} for user {target_id}")
        except Exception as e:
            logger.error(f"Failed to delete submission {sub_id}: {e}")
            await query.answer("❌ Failed to reject.", show_alert=True)
            return
        try:
            user_data = get_user(target_id)
            new_in_review = max(0, user_data.get("in_review", 0) - 1)
            update_user(target_id, {"in_review": new_in_review})
            try:
                await context.bot.send_message(
                    chat_id=target_id,
                    text=(
                        f"{DIV}\n"
                        "  ❌  Submission Rejected\n"
                        f"{DIV}\n\n"
                        "Your submission did not meet\n"
                        "our requirements.\n\n"
                        "Tips for approval:\n"
                        "  • Use a real mobile device\n"
                        "  • Enable 2FA correctly\n"
                        "  • Submit valid backup codes\n\n"
                        "Contact: @axWorker_Admin\n\n"
                        f"{DIV}"
                    ),
                    parse_mode=None
                )
            except Exception as e:
                logger.warning(f"Could not notify user {target_id}: {e}")
        except Exception as e:
            logger.error(f"Failed to update user {target_id}: {e}")
        cancelled_count = context.user_data.get("acts_cancelled_count", 0) + 1
        context.user_data["acts_cancelled_count"] = cancelled_count
        new_pending = [sub for sub in pending_subs if sub.get('id') != sub_id]
        context.user_data["acts_pending_subs"] = new_pending
        current_index = context.user_data.get("acts_current_index", 0)
        await _send_submission_for_review(update, context, current_index if new_pending else len(new_pending))
        await query.answer("❌ Rejected!")
        return


async def cmd_checktasks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    settings = get_task_settings()
    await update.message.reply_text(
        f"TASK SETTINGS\n\n"
        f"Standard 2FA: {'ENABLED' if settings.get('task_6h_enabled', True) else 'DISABLED'}\n"
        f"Premium 2FA:  {'ENABLED' if settings.get('task_1h_enabled', True) else 'DISABLED'}\n\n"
        f"Last updated: {settings.get('last_updated', 'Never')}"
    )


async def cmd_refreshtasks(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    await update.message.reply_text(
        "✅ Task menu refreshed.\n\n"
        "Users will see the updated task list."
    )


async def cmd_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    text = (
        "ADMIN COMMANDS\n"
        f"{DIV}\n\n"
        "USER MANAGEMENT\n"
        "/add {amount} {userid}\n"
        "/rm {amount} {userid}\n"
        "/apr {amount} {userid}\n"
        "/rmreview {userid}\n"
        "/userinfo {userid}\n"
        "/list\n"
        "/acts {userid}\n\n"
        "MESSAGING\n"
        "/msg {userid} {text}\n"
        "/cast {text}\n\n"
        "LEADERBOARD\n"
        "/ldset\n"
        "/ldauto\n"
        "/ldoff\n\n"
        "BOT CONTROL\n"
        "/botoff\n"
        "/boton\n\n"
        "TASK CONTROL\n"
        "/on2fa6h  /off2fa6h\n"
        "/on2fa1h  /off2fa1h\n\n"
        "SUBMISSIONS\n"
        "/rcv {userid}\n"
        "/rcvall\n"
        "/resetsub\n"
        "/live\n\n"
        "SETTINGS\n"
        "/stp {price}\n"
        "/stats\n"
        "/checktasks\n\n"
        "FACEBOOK COOKIE\n"
        "/fbon  /fboff\n"
        "/fbstp [amount]\n"
        "/fblive\n"
        "/fbrcv [userid]\n"
        "/fbrcvall\n"
        "/fbacts [userid]\n\n"
        f"{DIV}\n"
        "/cmd — this list"
    )
    await update.message.reply_text(text)

# ═════════════════════════════════════════════════════════════════════════════
# ADMIN PANEL — /panel
# ═════════════════════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────
# Panel Keyboard Builders
# ─────────────────────────────────────────────

def _panel_main_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("📊 Statistics",       callback_data="panel:stats"),
            InlineKeyboardButton("👥 Users",             callback_data="panel:users"),
        ],
        [
            InlineKeyboardButton("📋 IG Tasks",          callback_data="panel:ig_tasks"),
            InlineKeyboardButton("🍪 FB Tasks",          callback_data="panel:fb_tasks"),
        ],
        [
            InlineKeyboardButton("💰 Submissions",       callback_data="panel:submissions"),
            InlineKeyboardButton("🏆 Leaderboard",       callback_data="panel:leaderboard"),
        ],
        [
            InlineKeyboardButton("📢 Broadcast",         callback_data="panel:broadcast"),
            InlineKeyboardButton("🤖 Bot Control",       callback_data="panel:botcontrol"),
        ],
        [
            InlineKeyboardButton("💸 Total Withdrawals", callback_data="panel:withdrawals"),
        ],
        [
            InlineKeyboardButton("🔄 Refresh",           callback_data="panel:refresh"),
            InlineKeyboardButton("❌ Close",              callback_data="panel:close"),
        ],
    ])


def _panel_back_row() -> list:
    return [InlineKeyboardButton("⬅️ Back", callback_data="panel:main")]


def _panel_refresh_close_row() -> list:
    return [
        InlineKeyboardButton("🔄 Refresh", callback_data="panel:refresh"),
        InlineKeyboardButton("❌ Close",   callback_data="panel:close"),
    ]


def _panel_users_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("📋 User List",          callback_data="panel:list"),
            InlineKeyboardButton("🔍 Active Users (IG)",  callback_data="panel:live"),
        ],
        [
            InlineKeyboardButton("👤 User Info",          callback_data="panel:userinfo_prompt"),
            InlineKeyboardButton("➕ Add Balance",        callback_data="panel:add_prompt"),
        ],
        [
            InlineKeyboardButton("➖ Remove Balance",     callback_data="panel:rm_prompt"),
            InlineKeyboardButton("✅ Approve Tasks",      callback_data="panel:apr_prompt"),
        ],
        [
            InlineKeyboardButton("🗑️ Clear Review",      callback_data="panel:rmreview_prompt"),
            InlineKeyboardButton("📂 User Submissions",  callback_data="panel:acts_prompt"),
        ],
        _panel_back_row() + [InlineKeyboardButton("❌ Close", callback_data="panel:close")],
    ])


def _panel_ig_tasks_kb() -> InlineKeyboardMarkup:
    settings = get_task_settings()
    s6h = settings.get("task_6h_enabled", True)
    s1h = settings.get("task_1h_enabled", True)
    price = get_task_price()
    price_1h = get_task_1h_price()
    s6h_label = "✅ Standard ON" if s6h else "🔴 Standard OFF"
    s1h_label = "✅ Premium ON"  if s1h else "🔴 Premium OFF"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"📱 {s6h_label}",                   callback_data="panel:toggle_2fa6h")],
        [InlineKeyboardButton(f"⭐ {s1h_label}",                   callback_data="panel:toggle_2fa1h")],
        [InlineKeyboardButton(f"💲 IG Price: {fmt_usd(price)}",    callback_data="panel:stp_prompt")],
        [InlineKeyboardButton(f"🔐 Premium Price: {fmt_usd(price_1h)}", callback_data="panel:stp1h_prompt")],
        [InlineKeyboardButton("📊 Check Task Settings",             callback_data="panel:checktasks")],
        [InlineKeyboardButton("🔄 Refresh Tasks",                   callback_data="panel:refreshtasks")],
        _panel_back_row() + [InlineKeyboardButton("❌ Close", callback_data="panel:close")],
    ])


def _panel_fb_tasks_kb() -> InlineKeyboardMarkup:
    fb_settings = get_fb_task_settings()
    fb_on = fb_settings.get("fb_task_enabled", True)
    fb_price = get_fb_task_price()
    fb_label = "✅ FB Task ON" if fb_on else "🔴 FB Task OFF"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(f"🍪 {fb_label}",          callback_data="panel:toggle_fbtask")],
        [InlineKeyboardButton(f"💲 FB Price: {fmt_usd(fb_price)}", callback_data="panel:fbstp_prompt")],
        [InlineKeyboardButton("👀 FB Active Users",       callback_data="panel:fblive")],
        [InlineKeyboardButton("📂 FB Review",             callback_data="panel:fbacts_prompt")],
        [InlineKeyboardButton("📥 FB Download (User)",   callback_data="panel:fbrcv_prompt")],
        [InlineKeyboardButton("📥 FB Download (All)",    callback_data="panel:fbrcvall")],
        _panel_back_row() + [InlineKeyboardButton("❌ Close", callback_data="panel:close")],
    ])


def _panel_submissions_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📥 Download User XLSX",   callback_data="panel:rcv_prompt")],
        [InlineKeyboardButton("📥 Download All XLSX",    callback_data="panel:rcvall")],
        [InlineKeyboardButton("🗑️ Reset All Submissions",callback_data="panel:resetsub_confirm")],
        _panel_back_row() + [InlineKeyboardButton("❌ Close", callback_data="panel:close")],
    ])


def _panel_leaderboard_kb() -> InlineKeyboardMarkup:
    settings = get_leaderboard_settings()
    ld_on = settings.get("enabled", True)
    mode  = settings.get("mode", "auto")
    ld_label = f"✅ Leaderboard ON [{mode}]" if ld_on else "🔴 Leaderboard OFF"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(ld_label,                  callback_data="panel:ld_noop")],
        [InlineKeyboardButton("📊 Set Real Leaderboard", callback_data="panel:ldset")],
        [InlineKeyboardButton("🤖 Auto Leaderboard",     callback_data="panel:ldauto")],
        [InlineKeyboardButton("🔴 Disable Leaderboard",  callback_data="panel:ldoff")],
        _panel_back_row() + [InlineKeyboardButton("❌ Close", callback_data="panel:close")],
    ])


def _panel_broadcast_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📢 Broadcast to All",     callback_data="panel:cast_prompt")],
        [InlineKeyboardButton("✉️ Message a User",        callback_data="panel:msg_prompt")],
        _panel_back_row() + [InlineKeyboardButton("❌ Close", callback_data="panel:close")],
    ])


def _panel_botcontrol_kb() -> InlineKeyboardMarkup:
    bot_state = get_bot_state()
    enabled = bot_state.get("enabled", True)
    status_label = "🟢 Bot is ONLINE" if enabled else "🔴 Bot is OFFLINE"
    toggle_label = "🔴 Turn Bot OFF"   if enabled else "🟢 Turn Bot ON"
    toggle_cb    = "panel:botoff"       if enabled else "panel:boton"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(status_label,              callback_data="panel:bot_noop")],
        [InlineKeyboardButton(toggle_label,              callback_data=toggle_cb)],
        _panel_back_row() + [InlineKeyboardButton("❌ Close", callback_data="panel:close")],
    ])


def _panel_resetsub_confirm_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("⚠️ YES, Delete All",   callback_data="panel:resetsub_do"),
            InlineKeyboardButton("❌ Cancel",             callback_data="panel:submissions"),
        ],
    ])


# ─────────────────────────────────────────────
# Panel Main Text Builder
# ─────────────────────────────────────────────

def _panel_main_text() -> str:
    bot_state  = get_bot_state()
    enabled    = bot_state.get("enabled", True)
    bot_status = "🟢 ONLINE" if enabled else "🔴 OFFLINE"
    now        = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    try:
        all_users = get_all_users()
        total_users = len(all_users)
    except Exception:
        total_users = "?"
    try:
        all_subs   = get_all_submissions()
        pending_ig = sum(len(s) for s in all_subs.values() if s)
    except Exception:
        pending_ig = "?"
    try:
        fb_subs    = get_all_fb_submissions()
        pending_fb = sum(len(s) for s in fb_subs.values() if isinstance(s, dict))
    except Exception:
        pending_fb = "?"

    return (
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"  🛠  Admin Panel\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"Bot:          {bot_status}\n"
        f"Users:        {total_users}\n"
        f"IG Pending:   {pending_ig}\n"
        f"FB Pending:   {pending_fb}\n\n"
        f"Updated: {now}\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
    )


# ─────────────────────────────────────────────
# /panel command
# ─────────────────────────────────────────────

async def cmd_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Access Denied")
        return
    context.user_data.pop("panel_state", None)
    await update.message.reply_text(
        _panel_main_text(),
        parse_mode=None,
        reply_markup=_panel_main_kb(),
    )


async def _send_panel_withdrawal_page(query, context, idx: int):
    """Display one pending withdrawal request at a time with Prev/Next navigation."""
    pending_wds = context.user_data.get("panel_wd_list", [])
    total = len(pending_wds)
    if not pending_wds or idx >= total:
        await query.edit_message_text(
            "No pending withdrawal requests.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Back", callback_data="panel:main")]
            ]),
        )
        await query.answer()
        return

    wd = pending_wds[idx]
    user_id = wd.get("user_id", "?")
    w_id    = wd.get("w_id", "?")
    tg_user = wd.get("tg_username", "N/A")
    wallet  = wd.get("wallet", "N/A")
    amount  = wd.get("amount", 0.0)
    method  = wd.get("method", "usdt").upper()
    dt      = wd.get("datetime", "N/A")
    status  = wd.get("status", "pending").capitalize()
    short_wallet = f"{wallet[:10]}...{wallet[-6:]}" if len(wallet) > 16 else wallet

    text = (
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"  💸  Pending Withdrawal  {idx + 1}/{total}\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"Withdrawal ID:  {w_id}\n"
        f"User ID:        {user_id}\n"
        f"TG:             {tg_user}\n\n"
        f"Method:         {method}\n"
        f"Wallet:         {short_wallet}\n"
        f"Amount:         {fmt_usd(amount)}\n"
        f"Date:           {dt}\n"
        f"Status:         {status}\n\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
    )

    nav_row = []
    if idx > 0:
        nav_row.append(InlineKeyboardButton("⏪ Previous", callback_data=f"panel:panel_wd_page:{idx - 1}"))
    if idx < total - 1:
        nav_row.append(InlineKeyboardButton("⏭️ Next",     callback_data=f"panel:panel_wd_page:{idx + 1}"))

    keyboard = [
        [
            InlineKeyboardButton("✅ Success", callback_data=f"wd_approve:{user_id}:{w_id}"),
            InlineKeyboardButton("🚫 Cancel",  callback_data=f"wd_cancel:{user_id}:{w_id}"),
        ],
    ]
    if nav_row:
        keyboard.append(nav_row)
    keyboard.append([InlineKeyboardButton("⬅️ Back", callback_data="panel:main")])

    await query.edit_message_text(
        text,
        parse_mode=None,
        reply_markup=InlineKeyboardMarkup(keyboard),
    )
    await query.answer()


# ─────────────────────────────────────────────
# Panel Prompt Reply Helper
# ─────────────────────────────────────────────

async def _panel_send_prompt(query, context, prompt_text: str, state: str):
    """Edit the current panel message into a prompt and store state."""
    context.user_data["panel_state"] = state
    await query.edit_message_text(
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
        f"  📝  Input Required\n"
        f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
        f"{prompt_text}\n\n"
        f"Type /panel to cancel.",
        parse_mode=None,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("❌ Cancel", callback_data="panel:main")]
        ]),
    )
    await query.answer()


# ─────────────────────────────────────────────
# Panel Callback Handler
# ─────────────────────────────────────────────

async def callback_panel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query.from_user.id not in ADMIN_IDS:
        await query.answer("❌ Access Denied", show_alert=True)
        return

    data = query.data  # "panel:action"
    action = data.split(":", 1)[1] if ":" in data else data

    # ── No-op buttons (just labels) ──────────────────────────────────────────
    if action in ("bot_noop", "ld_noop"):
        await query.answer()
        return

    # ── Main panel ───────────────────────────────────────────────────────────
    if action in ("main", "refresh"):
        await query.edit_message_text(
            _panel_main_text(),
            parse_mode=None,
            reply_markup=_panel_main_kb(),
        )
        await query.answer("🔄 Refreshed" if action == "refresh" else "")
        return

    if action == "close":
        await query.edit_message_text("🛠 Admin Panel closed.")
        await query.answer()
        return

    # ── Sub-menus ────────────────────────────────────────────────────────────
    if action == "stats":
        await query.answer("⏳ Loading…")
        try:
            bot_state     = get_bot_state()
            bot_status    = "🟢 ONLINE" if bot_state.get("enabled", True) else "🔴 OFFLINE"
            all_users     = get_all_users()
            all_subs      = get_all_submissions()
            total_users   = len(all_users)
            total_subs    = sum(len(s) if isinstance(s, dict) else 0 for s in all_subs.values())
            total_approved = sum(u.get("approved", 0) for u in all_users.values())
            total_balance  = sum(u.get("balance",  0.0) for u in all_users.values())
            task_price     = get_task_price()
            today          = datetime.now(timezone.utc).strftime("%Y-%m-%d")
            today_subs     = sum(
                1 for subs in all_subs.values() if isinstance(subs, dict)
                for sub in subs.values() if isinstance(sub, dict) and sub.get("datetime","").startswith(today)
            )
            total_withdrawn = 0.0
            try:
                all_wds = db.reference("withdrawals").get() or {}
                for uid_wds in all_wds.values():
                    if isinstance(uid_wds, dict):
                        for w in uid_wds.values():
                            if isinstance(w, dict) and w.get("status") == "approved":
                                total_withdrawn += w.get("amount", 0.0)
            except Exception:
                pass
            fb_subs       = get_all_fb_submissions()
            fb_pending    = sum(len(s) for s in fb_subs.values() if isinstance(s, dict))
            text = (
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"  📊  Statistics\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                f"Bot Status:     {bot_status}\n"
                f"Since:          {bot_state.get('last_updated','?')}\n\n"
                f"USERS\n"
                f"Total:          {total_users}\n"
                f"Active:         {len([u for u in all_users.values() if u.get('total_submitted',0)>0])}\n\n"
                f"IG SUBMISSIONS\n"
                f"Total:          {total_subs}\n"
                f"Today:          {today_subs}\n"
                f"Approved:       {total_approved}\n\n"
                f"FB PENDING:     {fb_pending}\n\n"
                f"FINANCIALS\n"
                f"IG Price:       {fmt_usd(task_price)}\n"
                f"FB Price:       {fmt_usd(get_fb_task_price())}\n"
                f"Total Balance:  {fmt_usd(total_balance)}\n"
                f"Total Paid:     {fmt_usd(total_withdrawn)}\n\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━"
            )
        except Exception as e:
            text = f"❌ Error loading stats:\n{e}"
        await query.edit_message_text(
            text, parse_mode=None,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("🔄 Refresh", callback_data="panel:stats"),
                 InlineKeyboardButton("⬅️ Back",    callback_data="panel:main")],
            ]),
        )
        return

    if action == "users":
        await query.edit_message_text(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"  👥  User Management\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Select an action below.",
            parse_mode=None,
            reply_markup=_panel_users_kb(),
        )
        await query.answer()
        return

    if action == "ig_tasks":
        await query.edit_message_text(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"  📋  Instagram Task Control\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Toggle tasks or update pricing.",
            parse_mode=None,
            reply_markup=_panel_ig_tasks_kb(),
        )
        await query.answer()
        return

    if action == "fb_tasks":
        await query.edit_message_text(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"  🍪  Facebook Task Control\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Toggle FB task or manage submissions.",
            parse_mode=None,
            reply_markup=_panel_fb_tasks_kb(),
        )
        await query.answer()
        return

    if action == "submissions":
        all_subs  = get_all_submissions()
        pending   = sum(len(s) if isinstance(s, dict) else 0 for s in all_subs.values())
        await query.edit_message_text(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"  💰  Submission Management\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Pending IG submissions: {pending}\n\n"
            f"Download or reset submissions.",
            parse_mode=None,
            reply_markup=_panel_submissions_kb(),
        )
        await query.answer()
        return

    if action == "leaderboard":
        await query.edit_message_text(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"  🏆  Leaderboard Control\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Manage the public leaderboard.",
            parse_mode=None,
            reply_markup=_panel_leaderboard_kb(),
        )
        await query.answer()
        return

    if action == "broadcast":
        await query.edit_message_text(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"  📢  Broadcast & Messaging\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Send messages to users.",
            parse_mode=None,
            reply_markup=_panel_broadcast_kb(),
        )
        await query.answer()
        return

    if action == "botcontrol":
        await query.edit_message_text(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"  🤖  Bot Control\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Toggle the bot on or off.",
            parse_mode=None,
            reply_markup=_panel_botcontrol_kb(),
        )
        await query.answer()
        return

    # ── Bot ON / OFF ──────────────────────────────────────────────────────────
    if action == "botoff":
        set_bot_state(False, query.from_user.id)
        await query.answer("🔴 Bot disabled!")
        await query.edit_message_text(
            f"🔴 Bot is now OFFLINE.\n\nUse /panel → Bot Control to re-enable.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Back", callback_data="panel:botcontrol")]
            ]),
        )
        return

    if action == "boton":
        set_bot_state(True, query.from_user.id)
        await query.answer("🟢 Bot enabled!")
        await query.edit_message_text(
            f"🟢 Bot is now ONLINE.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Back", callback_data="panel:botcontrol")]
            ]),
        )
        return

    # ── IG Task Toggles ───────────────────────────────────────────────────────
    if action == "toggle_2fa6h":
        settings = get_task_settings()
        new_val  = not settings.get("task_6h_enabled", True)
        db.reference("settings/tasks/task_6h_enabled").set("true" if new_val else "false")
        db.reference("settings/tasks/last_updated").set(
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        )
        label = "ENABLED" if new_val else "DISABLED"
        await query.answer(f"📱 Standard 2FA {label}")
        await query.edit_message_text(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"  📋  Instagram Task Control\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Toggle tasks or update pricing.",
            parse_mode=None,
            reply_markup=_panel_ig_tasks_kb(),
        )
        return

    if action == "toggle_2fa1h":
        settings = get_task_settings()
        new_val  = not settings.get("task_1h_enabled", True)
        db.reference("settings/tasks/task_1h_enabled").set("true" if new_val else "false")
        db.reference("settings/tasks/last_updated").set(
            datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        )
        label = "ENABLED" if new_val else "DISABLED"
        await query.answer(f"⭐ Premium 2FA {label}")
        await query.edit_message_text(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"  📋  Instagram Task Control\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Toggle tasks or update pricing.",
            parse_mode=None,
            reply_markup=_panel_ig_tasks_kb(),
        )
        return

    if action == "checktasks":
        settings = get_task_settings()
        s6h = "ENABLED"  if settings.get("task_6h_enabled", True)  else "DISABLED"
        s1h = "ENABLED"  if settings.get("task_1h_enabled", True)  else "DISABLED"
        fb_s = get_fb_task_settings()
        fb  = "ENABLED"  if fb_s.get("fb_task_enabled", True)       else "DISABLED"
        await query.answer()
        await query.edit_message_text(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"  📊  Task Settings\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"IG Standard 2FA:  {s6h}\n"
            f"IG Premium 2FA:   {s1h}\n"
            f"FB Cookie:        {fb}\n\n"
            f"IG Price:  {fmt_usd(get_task_price())}\n"
            f"FB Price:  {fmt_usd(get_fb_task_price())}\n\n"
            f"Last updated: {settings.get('last_updated','Never')}\n\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
            parse_mode=None,
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Back", callback_data="panel:ig_tasks")]
            ]),
        )
        return

    if action == "refreshtasks":
        await query.answer("✅ Task menu refreshed for users.")
        await query.edit_message_text(
            f"✅ Task menu refreshed.\nUsers will see the updated task list.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Back", callback_data="panel:ig_tasks")]
            ]),
        )
        return

    # ── FB Task Toggle ────────────────────────────────────────────────────────
    if action == "toggle_fbtask":
        fb_s    = get_fb_task_settings()
        new_val = not fb_s.get("fb_task_enabled", True)
        set_fb_task_settings(new_val)
        label   = "ENABLED" if new_val else "DISABLED"
        await query.answer(f"🍪 FB Task {label}")
        await query.edit_message_text(
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
            f"  🍪  Facebook Task Control\n"
            f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
            f"Toggle FB task or manage submissions.",
            parse_mode=None,
            reply_markup=_panel_fb_tasks_kb(),
        )
        return

    # ── Leaderboard ───────────────────────────────────────────────────────────
    if action == "ldset":
        await query.answer("⏳ Generating…")
        try:
            lb  = generate_real_leaderboard()
            set_leaderboard_data(lb)
            set_leaderboard_settings({
                "mode": "real",
                "last_update": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                "enabled": True,
            })
            text = format_leaderboard_text(lb, "real")
            await query.edit_message_text(
                text, parse_mode=None,
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("⬅️ Back", callback_data="panel:leaderboard")]
                ]),
            )
        except Exception as e:
            await query.edit_message_text(f"❌ Error: {e}",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:leaderboard")]]))
        return

    if action == "ldauto":
        await query.answer("⏳ Generating…")
        try:
            previous = get_leaderboard_data()
            settings = get_leaderboard_settings()
            if settings.get("mode") == "auto" and previous:
                lb = generate_auto_leaderboard(increment=1, previous_leaderboard=previous)
                note = "Leaderboard incremented."
            else:
                lb = generate_auto_leaderboard(increment=0, previous_leaderboard=None)
                note = "Auto leaderboard generated (10 users)."
            set_leaderboard_data(lb)
            set_leaderboard_settings({
                "mode": "auto",
                "last_update": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                "enabled": True,
            })
            text = format_leaderboard_text(lb, "auto")
            await query.edit_message_text(
                f"✅ {note}\n\n" + text, parse_mode=None,
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("⬅️ Back", callback_data="panel:leaderboard")]
                ]),
            )
        except Exception as e:
            await query.edit_message_text(f"❌ Error: {e}",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:leaderboard")]]))
        return

    if action == "ldoff":
        set_leaderboard_settings({
            "enabled": False,
            "last_update": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        })
        await query.answer("🔴 Leaderboard disabled.")
        await query.edit_message_text(
            "🔴 Leaderboard is now OFF.\nUsers will see 'Leaderboard is offline'.",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("⬅️ Back", callback_data="panel:leaderboard")]
            ]),
        )
        return

    # ── Live User Lists ───────────────────────────────────────────────────────
    if action == "live":
        await query.answer("⏳ Loading…")
        try:
            all_users = get_all_users()
            all_subs  = get_all_submissions()
            active    = [
                {"user_id": int(uid), "submissions": len(all_subs.get(uid, {})), "balance": udata.get("balance", 0.0)}
                for uid, udata in all_users.items()
                if len(all_subs.get(uid, {})) > 0
            ]
            active.sort(key=lambda x: x["submissions"], reverse=True)
            if not active:
                await query.edit_message_text(
                    "No active IG users found.",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:users")]]),
                )
                return
            context.user_data["live_users"]       = active
            context.user_data["live_total_pages"] = (len(active) + 9) // 10
            await _send_live_page(update, context, 1)
        except Exception as e:
            await query.edit_message_text(f"❌ Error: {e}",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:users")]]))
        return

    if action == "list":
        await query.answer("⏳ Loading…")
        try:
            all_users = get_all_users()
            user_ids  = sorted([int(uid) for uid in all_users.keys()])
            if not user_ids:
                await query.edit_message_text(
                    "No users found.",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:users")]]),
                )
                return
            context.user_data["user_list"]   = user_ids
            context.user_data["total_pages"] = (len(user_ids) + 9) // 10
            await _send_user_list_page(update, context, 1)
        except Exception as e:
            await query.edit_message_text(f"❌ Error: {e}",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:users")]]))
        return

    if action == "fblive":
        await query.answer("⏳ Loading…")
        try:
            all_subs      = get_all_fb_submissions()
            all_fb_users  = get_all_fb_users()
            active = [
                {"user_id": int(uid), "submissions": len(subs), "balance": all_fb_users.get(uid, {}).get("balance", 0.0)}
                for uid, subs in all_subs.items()
                if isinstance(subs, dict) and subs
            ]
            active.sort(key=lambda x: x["submissions"], reverse=True)
            if not active:
                await query.edit_message_text(
                    "No active FB users found.",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:fb_tasks")]]),
                )
                return
            context.user_data["fb_live_users"]       = active
            context.user_data["fb_live_total_pages"] = (len(active) + 9) // 10
            await _send_fb_live_page(update, context, 1)
        except Exception as e:
            await query.edit_message_text(f"❌ Error: {e}",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:fb_tasks")]]))
        return

    # ── Submissions Download ──────────────────────────────────────────────────
    if action == "rcvall":
        await query.answer("⏳ Generating XLSX…")
        tmp_path = None
        try:
            all_subs = get_all_submissions()
            if not all_subs:
                await query.edit_message_text("No submissions found.",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:submissions")]]))
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Approved Submissions"
            ws.append(["#", "User ID", "TG Username", "IG Username", "Password", "2FA Key", "Date (UTC)", "Status"])
            row_num = 1
            for user_id_str, submissions in all_subs.items():
                if not isinstance(submissions, dict):
                    continue
                for sub_id, sub in submissions.items():
                    if not isinstance(sub, dict) or sub.get("status", "pending") != "approved":
                        continue
                    ws.append([row_num, sub.get("user_id", user_id_str), sub.get("tg_username",""),
                                sub.get("username",""), sub.get("password",""), sub.get("key",""),
                                sub.get("datetime",""), "approved"])
                    row_num += 1
            if row_num == 1:
                await query.edit_message_text("No approved submissions.",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:submissions")]]))
                return
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                wb.save(tmp.name)
                tmp_path = tmp.name
            with open(tmp_path, "rb") as f:
                await query.message.reply_document(document=f, filename="approved_submissions.xlsx",
                    caption=f"✅ Total approved: {row_num - 1}")
            await query.edit_message_text("✅ All approved submissions downloaded.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:submissions")]]))
        except Exception as e:
            await query.edit_message_text(f"❌ Error: {e}",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:submissions")]]))
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try: os.remove(tmp_path)
                except Exception: pass
        return

    if action == "fbrcvall":
        await query.answer("⏳ Generating FB XLSX…")
        tmp_path = None
        try:
            all_subs = get_all_fb_submissions()
            if not all_subs:
                await query.edit_message_text("No FB submissions found.",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:fb_tasks")]]))
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "All FB Submissions"
            ws.append(["#", "UID", "Password", "Cookies", "TG Username", "User ID", "DateTime"])
            row_num = 1
            for user_id_str, submissions in all_subs.items():
                if not isinstance(submissions, dict):
                    continue
                for sub_id, sub in submissions.items():
                    if not isinstance(sub, dict):
                        continue
                    ws.append([row_num, sub.get("uid",""), sub.get("password",""), sub.get("cookies",""),
                                sub.get("tg_username",""), sub.get("user_id", user_id_str), sub.get("datetime","")])
                    row_num += 1
            if row_num == 1:
                await query.edit_message_text("No FB submissions found.",
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:fb_tasks")]]))
                return
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                wb.save(tmp.name)
                tmp_path = tmp.name
            with open(tmp_path, "rb") as f:
                await query.message.reply_document(document=f, filename="all_fb_submissions.xlsx",
                    caption=f"Total FB submissions: {row_num - 1}")
            await query.edit_message_text("✅ All FB submissions downloaded.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:fb_tasks")]]))
        except Exception as e:
            await query.edit_message_text(f"❌ Error: {e}",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:fb_tasks")]]))
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try: os.remove(tmp_path)
                except Exception: pass
        return

    # ── Reset All Submissions ─────────────────────────────────────────────────
    if action == "resetsub_confirm":
        all_subs = get_all_submissions()
        total = sum(len(s) if isinstance(s, dict) else 0 for s in all_subs.values())
        await query.edit_message_text(
            f"⚠️ WARNING\n\n"
            f"This will permanently delete ALL {total} submission(s).\n\n"
            f"This action cannot be undone!\n\n"
            f"Are you sure?",
            reply_markup=_panel_resetsub_confirm_kb(),
        )
        await query.answer()
        return

    if action == "resetsub_do":
        await query.answer("⏳ Deleting…")
        try:
            all_subs = get_all_submissions()
            total = sum(len(s) if isinstance(s, dict) else 0 for s in all_subs.values())
            db.reference("submissions").delete()
            try: db.reference("xlsx_cache").delete()
            except Exception: pass
            all_users = db.reference("users").get() or {}
            for uid in all_users:
                try: db.reference(f"users/{uid}").update({"in_review": 0})
                except Exception: pass
            await query.edit_message_text(
                f"🗑️ All submissions deleted.\nRecords removed: {total}",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("⬅️ Back", callback_data="panel:submissions")]
                ]),
            )
            logger.info(f"Admin {query.from_user.id} reset all submissions ({total} records) via panel.")
        except Exception as e:
            await query.edit_message_text(f"❌ Error: {e}",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back", callback_data="panel:submissions")]]))
        return

    # ── Prompt-based actions (require text input) ─────────────────────────────
    PROMPTS = {
        "userinfo_prompt": ("Enter the User ID to look up:", "panel_userinfo"),
        "add_prompt":      ("Enter: amount userid\nExample: 0.5 123456789", "panel_add"),
        "rm_prompt":       ("Enter: amount userid\nExample: 0.5 123456789", "panel_rm"),
        "apr_prompt":      ("Enter: count userid\nExample: 3 123456789", "panel_apr"),
        "rmreview_prompt": ("Enter the User ID to clear review for:", "panel_rmreview"),
        "acts_prompt":     ("Enter the User ID to review submissions for:", "panel_acts"),
        "rcv_prompt":      ("Enter the User ID to download XLSX for:", "panel_rcv"),
        "fbrcv_prompt":    ("Enter the User ID to download FB XLSX for:", "panel_fbrcv"),
        "fbacts_prompt":   ("Enter the User ID to review FB submissions for:", "panel_fbacts"),
        "stp_prompt":      (f"Enter new IG task price.\nCurrent: {fmt_usd(get_task_price())}\nExample: 0.030", "panel_stp"),
        "stp1h_prompt":    (f"Enter new Instagram 2FA Premium task price.\nCurrent: {fmt_usd(get_task_1h_price())}\nExample: 0.220", "panel_stp1h"),
        "fbstp_prompt":    (f"Enter new FB task price.\nCurrent: {fmt_usd(get_fb_task_price())}\nExample: 0.035", "panel_fbstp"),
        "cast_prompt":     ("Enter the broadcast message to send to ALL users:", "panel_cast"),
        "msg_prompt":      ("Enter: userid message\nExample: 123456789 Hello!", "panel_msg"),
    }
    if action in PROMPTS:
        prompt_text, state = PROMPTS[action]
        await _panel_send_prompt(query, context, prompt_text, state)
        return

    # ── 💸 Total Withdrawals (paginated pending view) ─────────────────────────
    if action == "withdrawals":
        await query.answer("⏳ Loading…")
        pending_wds = get_all_pending_withdrawals()
        if not pending_wds:
            await query.edit_message_text(
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n"
                f"  💸  Total Withdrawals\n"
                f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
                "No pending withdrawal requests.\n\n"
                "━━━━━━━━━━━━━━━━━━━━━━━━━━━━",
                parse_mode=None,
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("⬅️ Back", callback_data="panel:main")]
                ]),
            )
            return
        context.user_data["panel_wd_list"] = pending_wds
        await _send_panel_withdrawal_page(query, context, 0)
        return

    if action.startswith("panel_wd_page:"):
        idx = int(action.split(":")[1])
        pending_wds = context.user_data.get("panel_wd_list")
        if not pending_wds:
            # Reload from DB if session lost
            pending_wds = get_all_pending_withdrawals()
            context.user_data["panel_wd_list"] = pending_wds
        if not pending_wds:
            await query.answer("No pending withdrawals.", show_alert=True)
            return
        await _send_panel_withdrawal_page(query, context, idx)
        return

    # Fallback
    await query.answer("⚠️ Unknown action.", show_alert=True)


# ─────────────────────────────────────────────
# Panel Text-Input Handler (for prompt states)
# ─────────────────────────────────────────────

async def panel_text_input_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handles text replies when the admin panel is awaiting input."""
    if not is_admin(update):
        return
    state = context.user_data.get("panel_state")
    if not state or not state.startswith("panel_"):
        return  # Not in panel input mode — let other handlers process

    text = update.message.text.strip()
    context.user_data.pop("panel_state", None)

    # Helper to send result with back button
    async def reply_ok(msg: str, back_cb: str):
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back to Panel", callback_data=back_cb)]])
        await update.message.reply_text(msg, parse_mode=None, reply_markup=kb)

    async def reply_err(msg: str, back_cb: str):
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back to Panel", callback_data=back_cb)]])
        await update.message.reply_text(f"❌ {msg}", parse_mode=None, reply_markup=kb)

    # ── /userinfo ─────────────────────────────────────────────────────────────
    if state == "panel_userinfo":
        try:
            user_id = int(text)
        except ValueError:
            await reply_err("Invalid user ID.", "panel:users")
            return
        try:
            data        = get_user(user_id)
            submissions = get_submissions(user_id)
            pending_wd  = 0
            total_wd    = 0.0
            try:
                withdrawals = db.reference(f"withdrawals/{user_id}").get()
                if withdrawals:
                    for w in withdrawals.values():
                        if isinstance(w, dict):
                            if w.get("status") == "pending":  pending_wd += 1
                            elif w.get("status") == "approved": total_wd += w.get("amount", 0.0)
            except Exception: pass
            tg_u = "N/A"
            if submissions:
                for s in submissions:
                    if s.get("tg_username"):
                        tg_u = s["tg_username"]; break
            earned = data.get("approved", 0) * get_task_price()
            await reply_ok(
                f"USER INFO\n━━━━━━━━━━━━━━━━━━\n\n"
                f"ID:         {user_id}\n"
                f"TG:         {tg_u}\n\n"
                f"Balance:    {fmt_usd(data.get('balance', 0.0))}\n"
                f"Withdrawn:  {fmt_usd(total_wd)}\n"
                f"Earned:     {fmt_usd(earned)}\n\n"
                f"Approved:   {data.get('approved', 0)}\n"
                f"In Review:  {data.get('in_review', 0)}\n"
                f"Submitted:  {data.get('total_submitted', 0)}\n"
                f"Pending:    {len(submissions)}\n"
                f"Pending WD: {pending_wd}\n\n"
                f"IG Price:   {fmt_usd(get_task_price())}",
                "panel:users",
            )
        except Exception as e:
            await reply_err(str(e), "panel:users")
        return

    # ── /add balance ──────────────────────────────────────────────────────────
    if state == "panel_add":
        parts = text.split()
        if len(parts) < 2:
            await reply_err("Format: amount userid", "panel:users"); return
        try:
            amount = float(parts[0]); target_id = int(parts[1])
        except ValueError:
            await reply_err("Invalid input.", "panel:users"); return
        try:
            ud = get_user(target_id)
            new_bal = round(ud.get("balance", 0.0) + amount, 4)
            update_user(target_id, {"balance": new_bal})
            await reply_ok(f"✅ Added {fmt_usd(amount)} to {target_id}.\nNew balance: {fmt_usd(new_bal)}", "panel:users")
        except Exception as e:
            await reply_err(str(e), "panel:users")
        return

    # ── /rm balance ───────────────────────────────────────────────────────────
    if state == "panel_rm":
        parts = text.split()
        if len(parts) < 2:
            await reply_err("Format: amount userid", "panel:users"); return
        try:
            amount = float(parts[0]); target_id = int(parts[1])
        except ValueError:
            await reply_err("Invalid input.", "panel:users"); return
        try:
            ud = get_user(target_id)
            new_bal = max(0.0, round(ud.get("balance", 0.0) - amount, 4))
            update_user(target_id, {"balance": new_bal})
            await reply_ok(f"✅ Removed {fmt_usd(amount)} from {target_id}.\nNew balance: {fmt_usd(new_bal)}", "panel:users")
        except Exception as e:
            await reply_err(str(e), "panel:users")
        return

    # ── /apr ──────────────────────────────────────────────────────────────────
    if state == "panel_apr":
        parts = text.split()
        if len(parts) < 2:
            await reply_err("Format: count userid", "panel:users"); return
        try:
            amount = int(parts[0]); target_id = int(parts[1])
        except ValueError:
            await reply_err("Invalid input.", "panel:users"); return
        try:
            ud = get_user(target_id)
            new_apr    = ud.get("approved", 0) + amount
            new_review = max(0, ud.get("in_review", 0) - amount)
            update_user(target_id, {"approved": new_apr, "in_review": new_review})
            await reply_ok(f"✅ +{amount} approved for {target_id}.\nTotal approved: {new_apr}", "panel:users")
        except Exception as e:
            await reply_err(str(e), "panel:users")
        return

    # ── /rmreview ─────────────────────────────────────────────────────────────
    if state == "panel_rmreview":
        try:
            target_id = int(text)
        except ValueError:
            await reply_err("Invalid user ID.", "panel:users"); return
        try:
            remove_submissions(target_id)
            update_user(target_id, {"in_review": 0})
            await reply_ok(f"✅ Submissions cleared for {target_id}.", "panel:users")
        except Exception as e:
            await reply_err(str(e), "panel:users")
        return

    # ── /acts (IG review) ─────────────────────────────────────────────────────
    if state == "panel_acts":
        try:
            target_id = int(text)
        except ValueError:
            await reply_err("Invalid user ID.", "panel:users"); return
        context.user_data["acts_target_id"] = target_id
        submissions = get_submissions(target_id)
        if not submissions:
            await reply_ok(f"No pending IG submissions for {target_id}.", "panel:users")
            return
        context.user_data["acts_pending_subs"]    = submissions
        context.user_data["acts_current_index"]   = 0
        context.user_data["acts_approved_count"]  = 0
        context.user_data["acts_cancelled_count"] = 0
        await _send_submission_for_review(update, context, 0)
        return

    # ── /rcv (IG download per user) ───────────────────────────────────────────
    if state == "panel_rcv":
        try:
            target_id = int(text)
        except ValueError:
            await reply_err("Invalid user ID.", "panel:submissions"); return
        subs = get_submissions(target_id)
        if not subs:
            await reply_ok(f"No IG submissions for user {target_id}.", "panel:submissions")
            return
        tmp_path = None
        try:
            xlsx_data = build_xlsx_bytes(target_id)
            if not xlsx_data:
                await reply_err("Could not generate XLSX.", "panel:submissions"); return
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(xlsx_data); tmp_path = tmp.name
            with open(tmp_path, "rb") as f:
                await update.message.reply_document(document=f,
                    filename=f"submissions_{target_id}.xlsx",
                    caption=f"IG Submissions for {target_id} ({len(subs)} records)")
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back to Panel", callback_data="panel:submissions")]])
            await update.message.reply_text("✅ Done.", reply_markup=kb)
        except Exception as e:
            await reply_err(str(e), "panel:submissions")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try: os.remove(tmp_path)
                except Exception: pass
        return

    # ── /fbrcv (FB download per user) ────────────────────────────────────────
    if state == "panel_fbrcv":
        try:
            target_id = int(text)
        except ValueError:
            await reply_err("Invalid user ID.", "panel:fb_tasks"); return
        subs = get_fb_submissions(target_id)
        if not subs:
            await reply_ok(f"No FB submissions for user {target_id}.", "panel:fb_tasks")
            return
        tmp_path = None
        try:
            xlsx_data = build_fb_xlsx_bytes(target_id)
            if not xlsx_data:
                await reply_err("Could not generate XLSX.", "panel:fb_tasks"); return
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(xlsx_data); tmp_path = tmp.name
            with open(tmp_path, "rb") as f:
                await update.message.reply_document(document=f,
                    filename=f"fb_submissions_{target_id}.xlsx",
                    caption=f"FB Submissions for {target_id} ({len(subs)} records)")
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back to Panel", callback_data="panel:fb_tasks")]])
            await update.message.reply_text("✅ Done.", reply_markup=kb)
        except Exception as e:
            await reply_err(str(e), "panel:fb_tasks")
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try: os.remove(tmp_path)
                except Exception: pass
        return

    # ── /fbacts (FB review) ───────────────────────────────────────────────────
    if state == "panel_fbacts":
        try:
            target_id = int(text)
        except ValueError:
            await reply_err("Invalid user ID.", "panel:fb_tasks"); return
        context.user_data["fbacts_target_id"] = target_id
        submissions = get_fb_submissions(target_id)
        if not submissions:
            await reply_ok(f"No pending FB submissions for {target_id}.", "panel:fb_tasks")
            return
        context.user_data["fbacts_pending_subs"]    = submissions
        context.user_data["fbacts_current_index"]   = 0
        context.user_data["fbacts_approved_count"]  = 0
        context.user_data["fbacts_cancelled_count"] = 0
        await _send_fb_submission_for_review(update, context, 0)
        return

    # ── /stp (IG price) ───────────────────────────────────────────────────────
    if state == "panel_stp":
        try:
            new_price = float(text)
            if new_price <= 0:
                await reply_err("Price must be > 0.", "panel:ig_tasks"); return
        except ValueError:
            await reply_err("Invalid price.", "panel:ig_tasks"); return
        old_price = get_task_price()
        set_task_price(new_price)
        await reply_ok(
            f"✅ IG task price updated.\nOld: {fmt_usd(old_price)}\nNew: {fmt_usd(new_price)}",
            "panel:ig_tasks",
        )
        return

    # ── Instagram 2FA Premium price ───────────────────────────────────────────
    if state == "panel_stp1h":
        try:
            new_price = float(text)
            if new_price <= 0:
                await reply_err("Price must be > 0.", "panel:ig_tasks"); return
        except ValueError:
            await reply_err("Invalid price.", "panel:ig_tasks"); return
        old_price = get_task_1h_price()
        set_task_1h_price(new_price)
        await reply_ok(
            f"✅ Instagram 2FA Premium price updated.\nOld: {fmt_usd(old_price)}\nNew: {fmt_usd(new_price)}",
            "panel:ig_tasks",
        )
        return

    # ── /fbstp (FB price) ─────────────────────────────────────────────────────
    if state == "panel_fbstp":
        try:
            new_price = float(text)
            if new_price <= 0:
                await reply_err("Price must be > 0.", "panel:fb_tasks"); return
        except ValueError:
            await reply_err("Invalid price.", "panel:fb_tasks"); return
        old_price = get_fb_task_price()
        set_fb_task_price(new_price)
        await reply_ok(
            f"✅ FB task price updated.\nOld: {fmt_usd(old_price)}\nNew: {fmt_usd(new_price)}",
            "panel:fb_tasks",
        )
        return

    # ── /cast (broadcast) ─────────────────────────────────────────────────────
    if state == "panel_cast":
        status_msg = await update.message.reply_text("📢 Broadcasting…")
        try:
            all_users = get_all_users()
            if not all_users:
                await status_msg.edit_text("❌ No users found.")
                return
            ok, fail = 0, 0
            for uid_str in all_users.keys():
                try:
                    await context.bot.send_message(chat_id=int(uid_str), text=text)
                    ok += 1
                except Exception:
                    fail += 1
            kb = InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ Back to Panel", callback_data="panel:broadcast")]])
            await status_msg.edit_text(
                f"✅ Broadcast done.\nSent: {ok}\nFailed: {fail}",
                reply_markup=kb,
            )
        except Exception as e:
            await status_msg.edit_text(f"❌ Error: {e}")
        return

    # ── /msg (DM a user) ──────────────────────────────────────────────────────
    if state == "panel_msg":
        parts = text.split(" ", 1)
        if len(parts) < 2:
            await reply_err("Format: userid message", "panel:broadcast"); return
        try:
            target_id = int(parts[0])
            msg_text  = parts[1]
        except ValueError:
            await reply_err("Invalid user ID.", "panel:broadcast"); return
        try:
            await context.bot.send_message(chat_id=target_id, text=msg_text)
            await reply_ok(f"✅ Message sent to {target_id}.", "panel:broadcast")
        except Exception as e:
            await reply_err(str(e), "panel:broadcast")
        return


# ─────────────────────────────────────────────
# Fallback
# ─────────────────────────────────────────────
async def handle_unknown(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    await update.message.reply_text(
        "Use the menu buttons below to navigate.",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME

# ─────────────────────────────────────────────
# Error Handler
# ─────────────────────────────────────────────
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.error("Exception occurred:", exc_info=context.error)
    if isinstance(update, Update) and update.effective_message:
        try:
            await update.effective_message.reply_text(
                "⚠️ Something went wrong. Please try again."
            )
        except Exception:
            pass

# ═════════════════════════════════════════════════════════════════════════════
# FACEBOOK COOKIE TASK SYSTEM
# ═════════════════════════════════════════════════════════════════════════════

_FB_FIRSTNAMES = [
    "James","John","Robert","Michael","William","David","Richard","Joseph","Thomas","Charles",
    "Emma","Olivia","Ava","Isabella","Sophia","Mia","Charlotte","Amelia","Harper","Evelyn",
    "Daniel","Matthew","Anthony","Mark","Donald","Steven","Paul","Andrew","Kenneth","Joshua",
    "Grace","Hannah","Lily","Zoe","Ella","Nora","Aria","Chloe","Penelope","Layla",
    "Ryan","Nathan","Aaron","Adam","Brian","Eric","Tyler","Jacob","Logan","Lucas",
]

_FB_LASTNAMES = [
    "Smith","Johnson","Williams","Brown","Jones","Garcia","Miller","Davis","Wilson","Moore",
    "Taylor","Anderson","Thomas","Jackson","White","Harris","Martin","Thompson","Young","King",
    "Walker","Allen","Scott","Adams","Baker","Carter","Mitchell","Nelson","Roberts","Turner",
    "Clark","Lewis","Robinson","Lee","Hall","Perez","Wright","Hill","Green","Evans",
    "Collins","Edwards","Stewart","Morris","Rogers","Reed","Bailey","Butler","Cox","Richardson",
]


def generate_fb_firstname() -> str:
    return random.choice(_FB_FIRSTNAMES)


def generate_fb_lastname() -> str:
    return random.choice(_FB_LASTNAMES)


def get_fb_task_price() -> float:
    try:
        price = db.reference("settings/fb_task_price").get()
        if price is None:
            price = 0.035
            db.reference("settings/fb_task_price").set(price)
        return float(price)
    except Exception as e:
        logger.error(f"get_fb_task_price failed: {e}")
        return 0.035


def set_fb_task_price(price: float):
    try:
        db.reference("settings/fb_task_price").set(round(price, 4))
    except Exception as e:
        logger.error(f"set_fb_task_price failed: {e}")


def get_fb_task_settings() -> dict:
    try:
        raw = db.reference("settings/fb_tasks/fb_task_enabled").get()
        def parse_bool(val, default=True):
            if val is None:
                return default
            if isinstance(val, bool):
                return val
            if isinstance(val, str):
                return val.lower() != "false"
            return bool(val)
        return {"fb_task_enabled": parse_bool(raw, True)}
    except Exception as e:
        logger.error(f"get_fb_task_settings failed: {e}")
        return {"fb_task_enabled": True}


def set_fb_task_settings(enabled: bool):
    try:
        db.reference("settings/fb_tasks/fb_task_enabled").set("true" if enabled else "false")
    except Exception as e:
        logger.error(f"set_fb_task_settings failed: {e}")


def get_fb_user(user_id: int) -> dict:
    try:
        ref = db.reference(f"fb_users/{user_id}")
        data = ref.get()
        if data is None:
            data = {"balance": 0.0, "approved": 0, "in_review": 0, "total_submitted": 0}
            ref.set(data)
        return data
    except Exception as e:
        logger.error(f"get_fb_user({user_id}) failed: {e}")
        return {"balance": 0.0, "approved": 0, "in_review": 0, "total_submitted": 0}


def update_fb_user(user_id: int, updates: dict):
    try:
        db.reference(f"fb_users/{user_id}").update(updates)
    except Exception as e:
        logger.error(f"update_fb_user({user_id}) failed: {e}")


def add_fb_submission(user_id: int, tg_username: str, uid: str, password: str, firstname: str, lastname: str):
    try:
        now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
        sub_ref = db.reference(f"fb_submissions/{user_id}").push()
        sub_ref.set({
            "uid": uid,
            "password": password,
            "firstname": firstname,
            "lastname": lastname,
            "cookies": "N/A",
            "tg_username": tg_username,
            "user_id": str(user_id),
            "datetime": now,
        })
        fb_user = get_fb_user(user_id)
        update_fb_user(user_id, {
            "in_review": fb_user.get("in_review", 0) + 1,
            "total_submitted": fb_user.get("total_submitted", 0) + 1,
        })
        _rebuild_fb_xlsx(user_id)
        logger.info(f"FB submission saved for user {user_id}, push_id={sub_ref.key}")
    except Exception as e:
        logger.error(f"add_fb_submission({user_id}) failed: {e}")


async def handle_task_fb_start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    firstname = generate_fb_firstname()
    lastname = generate_fb_lastname()
    password = get_default_password()
    context.user_data["fb_firstname"] = firstname
    context.user_data["fb_lastname"] = lastname
    context.user_data["fb_password"] = password

    await update.message.reply_text(
        f"First Name : <code>{firstname}</code>\n"
        f"Last Name  : <code>{lastname}</code>\n"
        f"Password   : <code>{password}</code>\n\n"
        f"Send Your <b>Email or Number</b>.",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup([["🔙 Cancel"]], resize_keyboard=True),
    )
    return TASK_FB_AWAIT_UID


async def handle_fb_await_uid(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    if text in ("🔙 Cancel", "Cancel ❌"):
        return await handle_fb_cancel(update, context)
    if not text:
        await update.message.reply_text(
            f"{DIV}\n"
            "  ℹ️  Email/Number Required\n"
            f"{DIV}\n\n"
            "Please send your Facebook account\n"
            "Email or Phone Number.\n\n"
            f"{DIV}"
        )
        return TASK_FB_AWAIT_UID
    
    # Save the email/number
    context.user_data["fb_uid"] = text
    logger.info(f"FB Email/Number received from user {update.effective_user.id}: {text}")
    
    # Ask if they created the account
    await update.message.reply_text(
        f"{DIV}\n"
        "  ✅  Email/Number Received\n"
        f"{DIV}\n\n"
        f"<code>{text}</code>\n\n"
        f"{DIV_SHORT}\n\n"
        "Are you sure you have created the\n"
        "Facebook account with the credentials\n"
        "provided above?\n\n"
        f"{DIV}",
        parse_mode="HTML",
        reply_markup=ReplyKeyboardMarkup(
            [
                ["✅ Account Registered"],
                ["❌ Cancel"],
            ],
            resize_keyboard=True,
            one_time_keyboard=True
        ),
    )
    return TASK_FB_STARTED
    
    
async def handle_fb_account_registered(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    email_or_number = context.user_data.get("fb_uid", "")
    firstname = context.user_data.get("fb_firstname", "")
    lastname = context.user_data.get("fb_lastname", "")
    password = context.user_data.get("fb_password", get_default_password())
    
    if not email_or_number:
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Session Expired\n"
            f"{DIV}\n\n"
            "Please start the task again.\n\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD
        )
        return HOME
    
    tg_username = f"@{user.username}" if user.username else str(user.id)
    
    try:
        # Use the updated function with firstname and lastname
        add_fb_submission(user.id, tg_username, email_or_number, password, firstname, lastname)
    except Exception as e:
        logger.error(f"handle_fb_account_registered failed: {e}")
        await update.message.reply_text(
            f"{DIV}\n"
            "  ⚠️  Submission Failed\n"
            f"{DIV}\n\n"
            "Could not save your submission.\n"
            "Please try again.\n\n"
            f"{DIV}",
            reply_markup=HOME_KEYBOARD,
        )
        return HOME

    context.user_data.pop("fb_uid", None)
    context.user_data.pop("fb_firstname", None)
    context.user_data.pop("fb_lastname", None)
    context.user_data.pop("fb_password", None)

    price = get_fb_task_price()
    
    await update.message.reply_text(
        f"{DIV}\n"
        "  ✅  Submission Received\n"
        f"{DIV}\n\n"
        "Your Facebook account has been\n"
        "saved for review.\n\n"
        f"Reward:        +{fmt_usd(price)}\n\n"
        "You'll be notified upon approval.\n\n"
        f"{DIV}",
        parse_mode=None,
        reply_markup=HOME_KEYBOARD,
    )
    return HOME
    
    
async def cmd_change(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if not context.args:
        await update.message.reply_text(
            "Usage: /Change {new_password}\n\n"
            "Example: /Change NewPass123!\n\n"
            "This will change the password for:\n"
            "• Instagram 2FA tasks\n"
            "• Facebook Cookie tasks"
        )
        return
    
    new_password = " ".join(context.args)
    
    try:
        db.reference("settings/default_password").set(new_password)
        await update.message.reply_text(
            f"✅ Default password updated!\n\n"
            f"New password: <code>{new_password}</code>\n\n"
            f"This will be used for:\n"
            f"• Instagram 2FA tasks\n"
            f"• Facebook Cookie tasks",
            parse_mode="HTML"
        )
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")
        

async def handle_fb_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.pop("fb_uid", None)
    context.user_data.pop("fb_firstname", None)
    context.user_data.pop("fb_lastname", None)
    context.user_data.pop("fb_password", None)
    await update.message.reply_text(
        f"{DIV}\n"
        "  ❌  Task Cancelled\n"
        f"{DIV}\n\n"
        "You can start a new task anytime\n"
        "from the Tasks menu.\n\n"
        f"{DIV}",
        reply_markup=HOME_KEYBOARD,
    )
    return HOME

# ─────────────────────────────────────────────
# FB Admin Commands
# ─────────────────────────────────────────────
async def cmd_fbon(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    set_fb_task_settings(True)
    await update.message.reply_text(
        f"✅ Facebook Cookie task ENABLED.\n"
        f"Reward: {fmt_usd(get_fb_task_price())}"
    )


async def cmd_fboff(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    set_fb_task_settings(False)
    await update.message.reply_text(
        "🔴 Facebook Cookie task DISABLED.\n"
        "Use /fbon to re-enable."
    )


async def cmd_fbstp(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if not context.args:
        current = get_fb_task_price()
        await update.message.reply_text(
            f"Current FB task price: {fmt_usd(current)}\n\n"
            "Usage: /fbstp 0.030"
        )
        return
    try:
        new_price = float(context.args[0])
        if new_price <= 0:
            await update.message.reply_text("❌ Price must be > 0.")
            return
        old_price = get_fb_task_price()
        set_fb_task_price(new_price)
        await update.message.reply_text(
            f"✅ FB price updated.\n\n"
            f"Old: {fmt_usd(old_price)}\n"
            f"New: {fmt_usd(new_price)}"
        )
    except ValueError:
        await update.message.reply_text("❌ Invalid price. Example: /fbstp 0.030")
    except Exception as e:
        await update.message.reply_text(f"❌ Error: {e}")


async def cmd_fblive(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("⏳ Fetching active FB users…")
    try:
        all_subs = get_all_fb_submissions()
        all_fb_users = get_all_fb_users()
        if not all_subs:
            await status_msg.edit_text("No Facebook submissions found.")
            return
        active_users = []
        for uid, subs in all_subs.items():
            if isinstance(subs, dict) and subs:
                fb_data = all_fb_users.get(uid, {})
                active_users.append({
                    "user_id": int(uid),
                    "submissions": len(subs),
                    "balance": fb_data.get("balance", 0.0),
                })
        active_users.sort(key=lambda x: x["submissions"], reverse=True)
        if not active_users:
            await status_msg.edit_text("No active FB users found.")
            return
        PAGE_SIZE = 10
        total_pages = (len(active_users) + PAGE_SIZE - 1) // PAGE_SIZE
        context.user_data["fb_live_users"] = active_users
        context.user_data["fb_live_total_pages"] = total_pages
        await _send_fb_live_page(update, context, 1)
    except Exception as e:
        logger.error(f"cmd_fblive failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")


async def _send_fb_live_page(update: Update, context: ContextTypes.DEFAULT_TYPE, page: int):
    active_users = context.user_data.get("fb_live_users", [])
    total_pages = context.user_data.get("fb_live_total_pages", 1)
    PAGE_SIZE = 10
    if not active_users:
        await update.message.reply_text("No FB active users found.")
        return
    start_idx = (page - 1) * PAGE_SIZE
    end_idx = start_idx + PAGE_SIZE
    page_users = active_users[start_idx:end_idx]
    lines = [
        f"FB ACTIVE USERS   Page {page}/{total_pages}",
        DIV_SHORT,
        "",
    ]
    for idx, u in enumerate(page_users, start=start_idx + 1):
        lines.append(
            f"{idx}. ID: {u['user_id']}\n"
            f"   Submissions: {u['submissions']}   "
            f"Balance: {fmt_usd(u['balance'])}\n"
        )
    lines.append(DIV_SHORT)
    message_text = "\n".join(lines)
    keyboard = []
    for u in page_users:
        keyboard.append([
            InlineKeyboardButton(
                f"📋 {u['user_id']}",
                callback_data=f"fb_live_copy_uid:{u['user_id']}",
            )
        ])
    nav = []
    if page > 1:
        nav.append(InlineKeyboardButton("◀ Prev", callback_data=f"fb_live_page:{page - 1}"))
    if page < total_pages:
        nav.append(InlineKeyboardButton("Next ▶", callback_data=f"fb_live_page:{page + 1}"))
    if nav:
        keyboard.append(nav)
    keyboard.append([InlineKeyboardButton("✖ Close", callback_data="fb_live_close")])
    reply_markup = InlineKeyboardMarkup(keyboard)
    if update.callback_query:
        await update.callback_query.edit_message_text(message_text, parse_mode=None, reply_markup=reply_markup)
        await update.callback_query.answer()
    else:
        await update.message.reply_text(message_text, parse_mode=None, reply_markup=reply_markup)


async def callback_fb_live_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    if query.from_user.id not in ADMIN_IDS:
        await query.answer("❌ Not authorised.", show_alert=True)
        return
    data = query.data
    if data.startswith("fb_live_page:"):
        page = int(data.split(":")[1])
        await _send_fb_live_page(update, context, page)
        return
    if data == "fb_live_close":
        await query.edit_message_text("FB live list closed.")
        await query.answer()
        return
    if data.startswith("fb_live_copy_uid:"):
        uid = data.split(":")[1]
        await query.answer("✅ See below!", show_alert=False)
        await query.message.reply_text(
            f"FB User ID:\n<code>{uid}</code>\n\n"
            f"Use with:\n"
            f"`/fbrcv {uid}` or `/fbacts {uid}`",
            parse_mode="HTML",
        )
        return


async def cmd_fbrcv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if not context.args:
        await update.message.reply_text("Usage: /fbrcv {userid}")
        return
    try:
        target_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID.")
        return
    subs = get_fb_submissions(target_id)
    if not subs:
        await update.message.reply_text(f"No FB submissions for user {target_id}.")
        return
    tmp_path = None
    try:
        xlsx_data = build_fb_xlsx_bytes(target_id)
        if not xlsx_data:
            await update.message.reply_text("❌ Could not generate XLSX.")
            return
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(xlsx_data)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename=f"fb_submissions_{target_id}.xlsx",
                caption=f"FB Submissions for {target_id} ({len(subs)} records)",
            )
    except Exception as e:
        logger.error(f"cmd_fbrcv failed: {e}")
        await update.message.reply_text(f"❌ Error: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


async def cmd_fbrcvall(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    status_msg = await update.message.reply_text("⏳ Generating FB XLSX...")
    tmp_path = None
    try:
        all_subs = get_all_fb_submissions()
        if not all_subs:
            await status_msg.edit_text("No FB submissions found.")
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "All FB Submissions"
        ws.append(["#", "UID", "Password", "Cookies", "TG Username", "User ID", "DateTime"])
        row_num = 1
        for user_id, submissions in all_subs.items():
            if not isinstance(submissions, dict):
                continue
            for sub_id, sub in submissions.items():
                if not isinstance(sub, dict):
                    continue
                ws.append([
                    row_num,
                    sub.get("uid", ""),
                    sub.get("password", ""),
                    sub.get("cookies", ""),
                    sub.get("tg_username", ""),
                    sub.get("user_id", str(user_id)),
                    sub.get("datetime", ""),
                ])
                row_num += 1
        if row_num == 1:
            await status_msg.edit_text("No FB submissions found.")
            return
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
        with open(tmp_path, "rb") as f:
            await update.message.reply_document(
                document=f,
                filename="all_fb_submissions.xlsx",
                caption=f"Total FB submissions: {row_num - 1}",
            )
        await status_msg.delete()
    except Exception as e:
        logger.error(f"cmd_fbrcvall failed: {e}")
        await status_msg.edit_text(f"❌ Error: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


async def cmd_fbacts(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("❌ Unauthorised.")
        return
    if not context.args:
        await update.message.reply_text("Usage: /fbacts {userid}")
        return
    try:
        target_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ Invalid user ID.")
        return
    context.user_data["fbacts_target_id"] = target_id
    submissions = get_fb_submissions(target_id)
    if not submissions:
        await update.message.reply_text(f"No pending FB submissions for user {target_id}.")
        return
    context.user_data["fbacts_pending_subs"] = submissions
    context.user_data["fbacts_current_index"] = 0
    context.user_data["fbacts_approved_count"] = 0
    context.user_data["fbacts_cancelled_count"] = 0
    await _send_fb_submission_for_review(update, context, 0)
    return ADMIN_FBACTS_VIEW


async def _send_fb_submission_for_review(update: Update, context: ContextTypes.DEFAULT_TYPE, index: int):
    pending_subs = context.user_data.get("fbacts_pending_subs", [])
    target_id = context.user_data.get("fbacts_target_id")
    
    if not pending_subs or index >= len(pending_subs):
        approved_count = context.user_data.get("fbacts_approved_count", 0)
        cancelled_count = context.user_data.get("fbacts_cancelled_count", 0)
        total = len(context.user_data.get("fbacts_pending_subs", [])) + approved_count + cancelled_count
        
        summary = (
            f"✅ FB Review Complete!\n\n"
            f"Total Accounts:  {total}\n"
            f"Approved:  {approved_count}\n"
            f"Rejected:  {cancelled_count}"
        )
        for key in ("fbacts_pending_subs", "fbacts_target_id", "fbacts_current_index",
                    "fbacts_approved_count", "fbacts_cancelled_count"):
            context.user_data.pop(key, None)
        if update.callback_query:
            await update.callback_query.edit_message_text(summary, parse_mode=None)
            await update.callback_query.answer()
        else:
            await update.message.reply_text(summary, parse_mode=None)
        return

    sub = pending_subs[index]
    total = len(pending_subs)
    current = index + 1
    
    email_number = sub.get("uid", "N/A")
    firstname = sub.get("firstname", "N/A")
    lastname = sub.get("lastname", "N/A")
    password = sub.get("password", "N/A")
    
    submission_text = (
        f"📱 FB ACCOUNT REVIEW   [{current}/{total}]\n"
        f"{DIV_SHORT}\n"
        f"User ID:     {target_id}\n\n"
        f"First Name:  <code>{firstname}</code>\n"
        f"Last Name:   <code>{lastname}</code>\n"
        f"Email/Phone: <code>{email_number}</code>\n"
        f"Password:    <code>{password}</code>\n\n"
        f"TG: {sub.get('tg_username', 'N/A')}\n"
        f"At: {sub.get('datetime', 'N/A')}"
    )
    
    sub_id = sub.get("id")
    keyboard = [
        [
            InlineKeyboardButton("✅ Approve", callback_data=f"fbacts_approve:{target_id}:{sub_id}:{index}"),
            InlineKeyboardButton("❌ Reject", callback_data=f"fbacts_cancel:{target_id}:{sub_id}:{index}"),
        ],
        [
            InlineKeyboardButton("◀️ Previous", callback_data=f"fbacts_prev:{index}"),
            InlineKeyboardButton("Next ▶️", callback_data=f"fbacts_next:{index}"),
        ],
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    
    if update.callback_query:
        await update.callback_query.edit_message_text(
            submission_text, 
            parse_mode="HTML", 
            reply_markup=reply_markup
        )
        await update.callback_query.answer()
    else:
        await update.message.reply_text(
            submission_text, 
            parse_mode="HTML", 
            reply_markup=reply_markup
        )


async def callback_fb_copy_user_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    data = query.data  # fb_copy:fieldname:value
    parts = data.split(":", 2)
    if len(parts) < 3:
        await query.answer("⚠️ Invalid callback.", show_alert=True)
        return
    field = parts[1]
    value = parts[2]
    label_map = {"firstname": "First Name", "lastname": "Last Name", "password": "Password"}
    label = label_map.get(field, field.capitalize())
    await query.answer("✅ See below!", show_alert=False)
    await query.message.reply_text(
        f"{label}:\n<code>{value}</code>\n\nPress and hold to copy.",
        parse_mode="HTML",
    )

# ═════════════════════════════════════════════════════════════════════════════
# Main
# ═════════════════════════════════════════════════════════════════════════════
def main():
    init_firebase()

    app = Application.builder().token(BOT_TOKEN).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", cmd_start)],
        states={
    HOME: [
        MessageHandler(filters.Regex("^💼 Dashboard$"), handle_dashboard),
        MessageHandler(filters.Regex("^💰 Wallet$"), handle_balance),
        MessageHandler(filters.Regex("^📋 Tasks$"), handle_tasks),
        MessageHandler(filters.Regex("^🏆 Leaderboard$"), handle_leaderboard),
        MessageHandler(filters.Regex("^👥 Invite Friends$"), handle_referrals),
        MessageHandler(filters.Regex("^📥 Withdraw$"), handle_withdraw),
        MessageHandler(filters.Regex("^👤 Profile$"), handle_profile),
        MessageHandler(filters.Regex("^🫟 I'm New User$"), handle_new_user),
        MessageHandler(filters.Regex("^📞 Support$"), handle_support),
        MessageHandler(filters.Regex("^⚙️ Settings$"), handle_settings),
    ],
    SETTINGS_MENU: [
        MessageHandler(filters.Regex("^ℹ️ About Bot$"), handle_about_bot),
        MessageHandler(filters.Regex("^🔙 Back$"), handle_back_to_home),
    ],
    TASK_MENU: [
        MessageHandler(filters.Regex(r"^📱 Instagram 2FA — \$[\d.]+$"), handle_task_2fa_info),
        MessageHandler(filters.Regex(r"^⭐ Instagram 2FA Premium — \$[\d.]+$"), handle_task_2fa_1h_info),
        MessageHandler(filters.Regex(r"^🍪 Facebook Cookie — \$[\d.]+$"), handle_task_fb_info),
        MessageHandler(filters.Regex("^🔙 Back$"), handle_back_to_home),
    ],
    TASK_FB_INFO: [
        MessageHandler(filters.Regex("^▶️ Start Task$"), handle_task_fb_start),
        MessageHandler(filters.Regex("^🔙 Cancel$"), handle_fb_cancel),
        MessageHandler(filters.Regex("^🔙 Back$"), handle_back_to_home),
    ],
    # ⭐⭐⭐ এই দুইটা ADD করো ⭐⭐⭐
    TASK_FB_AWAIT_UID: [
        MessageHandler(filters.Regex("^🔙 Cancel$"), handle_fb_cancel),
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_fb_await_uid),
    ],
    TASK_FB_STARTED: [
        MessageHandler(filters.Regex("^✅ Account Registered$"), handle_fb_account_registered),
        MessageHandler(filters.Regex("^❌ Cancel$"), handle_fb_cancel),
    ],
    # ⭐⭐⭐ এখান পর্যন্ত ⭐⭐⭐
    TASK_2FA_INFO: [
        MessageHandler(filters.Regex("^🔙 Cancel$"), handle_task_cancel),
        MessageHandler(filters.Regex("^▶️ Start Task$"), handle_task_start),
        MessageHandler(filters.Regex("^🔙 Back$"), handle_back_to_home),
    ],
    TASK_2FA_STARTED: [
        MessageHandler(filters.Regex("^✅ Account Registered$"), handle_account_registered),
        MessageHandler(filters.Regex("^❌ Cancel Task$"), handle_2fa_cancel),
    ],
    TASK_2FA_AWAIT_KEY: [
        MessageHandler(filters.Regex("^🔙 Cancel$"), handle_task_cancel),
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_2fa_key),
    ],
    TASK_2FA_1H_INFO: [
        MessageHandler(filters.Regex("^🔙 Cancel$"), handle_task_1h_cancel),
        MessageHandler(filters.Regex("^▶️ Start Task$"), handle_task_1h_start),
        MessageHandler(filters.Regex("^🔙 Back$"), handle_back_to_home),
    ],
    TASK_2FA_1H_AWAIT_KEY: [
        MessageHandler(filters.Regex("^🔙 Cancel$"), handle_task_1h_cancel),
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_2fa_key_1h),
    ],
    TASK_2FA_1H_STARTED: [
        MessageHandler(filters.Regex("^✅ Account Registered$"), handle_account_registered_1h),
        MessageHandler(filters.Regex("^❌ Cancel Task$"), handle_task_1h_cancel),
    ],
    WITHDRAW_MENU: [
        MessageHandler(filters.Regex(r"^💎 USDT-BEP20$"), handle_withdraw_new_usdt),
        MessageHandler(filters.Regex(r"^💎 USDT — BEP20$"), handle_withdraw_bep20),
        MessageHandler(filters.Regex(r"^📱 bKash — BDT$"), handle_withdraw_bkash),
        MessageHandler(filters.Regex("^Cancel ❌$"), handle_withdraw_new_cancel),
        MessageHandler(filters.Regex("^🔙 Back$"), handle_back_to_home),
    ],
    WITHDRAW_CONFIRM: [
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_back_to_home),
    ],
    WITHDRAW_AMOUNT: [
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_withdraw_amount),
    ],
    WITHDRAW_ADDRESS: [
        MessageHandler(filters.TEXT & ~filters.COMMAND, handle_withdraw_address),
    ],
    },

    app.add_handler(conv)

    # ── Admin Panel ──────────────────────────────────────────────────────────
    app.add_handler(CommandHandler("panel", cmd_panel))
    app.add_handler(CallbackQueryHandler(callback_panel, pattern="^panel:"))
    # Panel text input handler (group=1 runs alongside ConversationHandler without interfering)
    app.add_handler(MessageHandler(
        filters.TEXT & ~filters.COMMAND & filters.User(list(ADMIN_IDS)),
        panel_text_input_handler,
    ), group=1)

    # Callback handlers
    app.add_handler(CallbackQueryHandler(callback_withdrawal, pattern="^wd_(approve|cancel):"))
    app.add_handler(CallbackQueryHandler(callback_withdraw_confirm, pattern="^wd_confirm"))
    app.add_handler(CallbackQueryHandler(callback_live_handler, pattern="^(live_page:|live_copy_uid:|live_close)"))
    app.add_handler(CallbackQueryHandler(callback_check_join, pattern="^check_join$"))
    app.add_handler(CallbackQueryHandler(callback_acts_copy_handler, pattern="^(acts_copy_username:|acts_copy_password:|acts_copy_2fa:)"))
    app.add_handler(CallbackQueryHandler(callback_2fa_handler, pattern="^(copy_otp:|confirm_registered|cancel_2fa_task)"))
    app.add_handler(CallbackQueryHandler(callback_list_handlers, pattern="^(list_page:|copy_uid:|list_close)"))
    app.add_handler(CallbackQueryHandler(callback_quick_review, pattern="^(quick_approve:|quick_cancel:)"))
    app.add_handler(CallbackQueryHandler(callback_acts_handler, pattern="^(acts_approve:|acts_cancel:|acts_exit)"))
    # FB callbacks
    app.add_handler(CallbackQueryHandler(callback_fb_copy_user_handler, pattern="^fb_copy:"))
    app.add_handler(CallbackQueryHandler(callback_fb_live_handler, pattern="^(fb_live_page:|fb_live_copy_uid:|fb_live_close)"))
    app.add_handler(CallbackQueryHandler(callback_fb_copy_handler, pattern="^(fbacts_copy_uid:|fbacts_copy_password:|fbacts_copy_cookies:)"))
    app.add_handler(CallbackQueryHandler(callback_fb_acts_handler, pattern="^(fbacts_approve:|fbacts_cancel:|fbacts_exit)"))

    # Admin commands
    app.add_handler(CommandHandler("live", cmd_live))
    app.add_handler(CommandHandler("rcv", cmd_rcv))
    app.add_handler(CommandHandler("rcvall", cmd_rcvall))
    app.add_handler(CommandHandler("resetsub", cmd_resetsub))
    app.add_handler(CommandHandler("add", cmd_add))
    app.add_handler(CommandHandler("rm", cmd_rm))
    app.add_handler(CommandHandler("msg", cmd_msg))
    app.add_handler(CommandHandler("cast", cmd_cast))
    app.add_handler(CommandHandler("userinfo", cmd_userinfo))
    app.add_handler(CommandHandler("stp", cmd_stp))
    app.add_handler(CommandHandler("ldset", cmd_ldset))
    app.add_handler(CommandHandler("ldauto", cmd_ldauto))
    app.add_handler(CommandHandler("ldoff", cmd_ldoff))
    app.add_handler(CommandHandler("acts", cmd_acts))
    app.add_handler(CommandHandler("list", cmd_list))
    app.add_handler(CommandHandler("checktasks", cmd_checktasks))
    app.add_handler(CommandHandler("refreshtasks", cmd_refreshtasks))
    app.add_handler(CommandHandler("botoff", cmd_botoff))
    app.add_handler(CommandHandler("Change", cmd_change))
    app.add_handler(CommandHandler("boton", cmd_boton))
    app.add_handler(CommandHandler("on2fa6h", cmd_on2fa6h))
    app.add_handler(CommandHandler("off2fa6h", cmd_off2fa6h))
    app.add_handler(CommandHandler("on2fa1h", cmd_on2fa1h))
    app.add_handler(CommandHandler("off2fa1h", cmd_off2fa1h))
    app.add_handler(CommandHandler("stats", cmd_stats))
    app.add_handler(CommandHandler("rmreview", cmd_rmreview))
    app.add_handler(CommandHandler("apr", cmd_apr))
    app.add_handler(CommandHandler("cmd", cmd_cmd))
    # FB commands
    app.add_handler(CommandHandler("fbon", cmd_fbon))
    app.add_handler(CommandHandler("fboff", cmd_fboff))
    app.add_handler(CommandHandler("fbstp", cmd_fbstp))
    app.add_handler(CommandHandler("fblive", cmd_fblive))
    app.add_handler(CommandHandler("fbrcv", cmd_fbrcv))
    app.add_handler(CommandHandler("fbrcvall", cmd_fbrcvall))
    app.add_handler(CommandHandler("fbacts", cmd_fbacts))

    app.add_error_handler(error_handler)

    logger.info("Bot is starting (polling)…")
    app.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)


if __name__ == "__main__":
    main()
